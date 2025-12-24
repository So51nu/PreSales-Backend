from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from django.db import transaction
import json
from .models import Inventory
from .serializers_inventory import InventorySerializer,InventoryDetailSerializer
import pandas as pd
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from .models import UnitType,UnitConfiguration,Facing

import json
import pandas as pd

from setup.models import UnitType, UnitConfiguration, Facing  # jahan ye models hain
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from django.db import transaction

# class InventoryViewSet(viewsets.ModelViewSet):
#     """
#     Endpoints:
#       - GET    /api/client/inventory/?project_id=...
#       - POST   /api/client/inventory/               (single create; supports multipart for photo + docs[file])
#       - POST   /api/client/inventory/bulk-create/   (bulk create; JSON only)
#       - PATCH  /api/client/inventory/{id}/
#       - PUT    /api/client/inventory/{id}/
#       - GET    /api/client/inventory/{id}/
#       - DELETE /api/client/inventory/{id}/
#     """
#     queryset = Inventory.objects.select_related(
#         "project", "tower", "floor", "unit",
#         "unit_type", "configuration", "facing"
#     ).all()
#     serializer_class = InventorySerializer
#     permission_classes = [IsAuthenticated]
#     parser_classes = (JSONParser, MultiPartParser, FormParser)

#     def get_queryset(self):
#         qs = super().get_queryset()
#         q = self.request.query_params

#         project_id = q.get("project_id")
#         if project_id:
#             qs = qs.filter(project_id=project_id)

#         availability = q.get("availability_status")
#         if availability:
#             qs = qs.filter(availability_status=availability)

#         status_val = q.get("status")
#         if status_val:
#             qs = qs.filter(status=status_val)

#         unit_status = q.get("unit_status")
#         if unit_status:
#             qs = qs.filter(unit_status=unit_status)

#         tower_id = q.get("tower_id")
#         if tower_id:
#             qs = qs.filter(tower_id=tower_id)

#         return qs.order_by("-id")


#     @action(detail=False, methods=["post"], url_path="bulk-create")
#     def bulk_create(self, request):
#         """
#         Bulk create inventory.

#         Supports:
#         1) JSON array body
#         2) multipart with "items" JSON
#         3) Excel file upload: multipart with "file" = .xlsx
#         """
#         raw = request.data

#         excel_file = request.FILES.get("file") or request.FILES.get("excel")
#         if excel_file:
#             # ---- CASE 3: Excel upload ----
#             try:
#                 df = pd.read_excel(excel_file)  # needs openpyxl
#             except Exception as e:
#                 return Response(
#                     {"detail": f"Failed to read Excel: {e}"},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )

#             # 1) Excel headers -> serializer field names
#             #    (ye tumhari current file ke column names hai)
#             rename_map = {
#                 "project_id": "project",
#                 "tower_id": "tower",
#                 "floor_id": "floor",
#                 "unit_id": "unit",

#                 "carpet_area": "carpet_sqft",
#                 "saleable_area": "saleable_sqft",


#     # NEW mappings (change left keys to EXACT Excel column names)
#     "Core Base Price (psf)": "core_base_price_psf",
#     "Approved Limit Price (psf)": "approved_limit_price_psf",
#     "Customer Base Price (psf)": "customer_base_price_psf",

#                 "base_rate": "rate_psf",
#                 "base_amount": "agreement_value",

#                 "remarks": "description",
#                 # agar future me "block_days" column add karo:
#                 "block_days": "block_period_days",
#             }
#             df.rename(columns=rename_map, inplace=True)

#             # 2) NaN -> None
#             df = df.where(df.notnull(), None)

#             # 3) Row-wise cleanup + code -> id mapping
#             rows = []
#             for row in df.to_dict(orient="records"):

#                 # a) FK IDs: project/tower/floor/unit/unit_type/configuration/facing
#                 for key in [
#                     "project",
#                     "tower",
#                     "floor",
#                     "unit",
#                     "unit_type",
#                     "configuration",
#                     "facing",
#                 ]:
#                     val = row.get(key)
#                     if val in ("", None):
#                         row[key] = None
#                     else:
#                         try:
#                             row[key] = int(val)
#                         except (ValueError, TypeError):
#                             # agar int nahi bana to rehne do; serializer handle karega
#                             pass

#                 # b) unit_type_code -> unit_type id (optional)
#                 utc = row.get("unit_type_code")
#                 if utc and not row.get("unit_type"):
#                     ut = UnitType.objects.filter(code=str(utc).upper()).first()
#                     if ut:
#                         row["unit_type"] = ut.id

#                 # c) configuration_code -> configuration id (optional)
#                 ucc = row.get("configuration_code")
#                 if ucc and not row.get("configuration"):
#                     cfg = UnitConfiguration.objects.filter(
#                         code=str(ucc).upper()
#                     ).first()
#                     if cfg:
#                         row["configuration"] = cfg.id

#                 # d) facing_code -> facing id (optional)
#                 fc = row.get("facing_code")
#                 if fc and not row.get("facing"):
#                     fobj = Facing.objects.filter(code=str(fc).upper()).first()
#                     if fobj:
#                         row["facing"] = fobj.id

#                 rows.append(row)

#             items = rows
#             files_map = {}  # Excel me abhi docs nahi

#         elif isinstance(raw, list):
#             # ---- CASE 1: pure JSON array ----
#             items = raw
#             files_map = request.FILES

#         else:
#             # ---- CASE 2: multipart with items JSON ----
#             items_json = raw.get("items")
#             if not items_json:
#                 return Response(
#                     {"detail": 'Expected a JSON list, "items" field, or Excel "file".'},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )
#             try:
#                 items = json.loads(items_json)
#             except json.JSONDecodeError:
#                 return Response(
#                     {"detail": 'Invalid JSON in "items".'},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )
#             files_map = request.FILES

#         if not isinstance(items, list):
#             return Response({"detail": "Items must be a list."}, status=400)

#         created = []
#         errors = []

#         with transaction.atomic():
#             for idx, payload in enumerate(items, start=1):
#                 payload = dict(payload)

#                 # documents (React path jaisa)
#                 docs_meta = payload.pop("documents", [])
#                 resolved_docs = []

#                 for doc in docs_meta or []:
#                     file_key = doc.get("file_field")
#                     file_obj = files_map.get(file_key) if file_key else None
#                     if not file_obj:
#                         continue
#                     doc_data = dict(doc)
#                     doc_data.pop("file_field", None)
#                     doc_data["file"] = file_obj
#                     resolved_docs.append(doc_data)

#                 if resolved_docs:
#                     payload["documents"] = resolved_docs

#                 ser = InventorySerializer(data=payload, context={"request": request})
#                 if ser.is_valid():
#                     inv = ser.save()
#                     created.append(inv.id)
#                 else:
#                     errors.append({"index": idx, "errors": ser.errors})

#         code = (
#             status.HTTP_201_CREATED
#             if created and not errors
#             else status.HTTP_207_MULTI_STATUS
#         )
#         return Response({"created_ids": created, "errors": errors}, status=code)


#     @action(detail=False, methods=["get"], url_path="by-unit")
#     def by_unit(self, request):
#         """
#         GET /api/client/inventory/by-unit/?unit_id=<id>

#         Returns Inventory of that Unit with:
#           - FK names (project_name, tower_name, floor_number, unit_no, etc.)
#           - documents with type labels
#         """
#         unit_id = request.query_params.get("unit_id")
#         if not unit_id:
#             return Response(
#                 {"detail": "unit_id is required."},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         inv = get_object_or_404(
#             Inventory.objects.select_related(
#                 "project",
#                 "tower",
#                 "floor",
#                 "unit",
#                 "unit_type",
#                 "configuration",
#                 "facing",
#             ).prefetch_related("documents"),
#             unit_id=unit_id,
#         )

#         serializer = InventoryDetailSerializer(
#             inv, context={"request": request}
#         )
#         return Response(serializer.data, status=status.HTTP_200_OK)

class InventoryViewSet(viewsets.ModelViewSet):
    """
    Endpoints:
      - GET    /api/client/inventory/?project_id=...
      - POST   /api/client/inventory/
      - POST   /api/client/inventory/bulk-create/
      - PATCH  /api/client/inventory/{id}/
      - PUT    /api/client/inventory/{id}/
      - GET    /api/client/inventory/{id}/
      - DELETE /api/client/inventory/{id}/
    """

    queryset = Inventory.objects.select_related(
        "project", "tower", "floor", "unit",
        "unit_type", "configuration", "facing"
    )
    serializer_class = InventorySerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (JSONParser, MultiPartParser, FormParser)

    # =====================================================
    # QUERYSET (ADMIN + FULL_CONTROL FIXED)
    # =====================================================
    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        q = self.request.query_params

        # 🔐 OWNERSHIP FILTER (INLINE)
        if not user.is_staff:
            role = getattr(user, "role", None)
            role_code = role.upper() if isinstance(role, str) else getattr(role, "code", None)

            if role_code == "ADMIN":
                qs = qs.filter(project__belongs_to=user)

            elif role_code == "FULL_CONTROL":
                qs = qs.filter(project__belongs_to_id=getattr(user, "admin_id", None))

            else:  # SALES / RECEPTION / OTHERS
                qs = qs.filter(project__belongs_to_id=getattr(user, "admin_id", None))

        # ---- filters ----
        project_id = q.get("project_id")
        if project_id:
            qs = qs.filter(project_id=project_id)

        availability = q.get("availability_status")
        if availability:
            qs = qs.filter(availability_status=availability)

        status_val = q.get("status")
        if status_val:
            qs = qs.filter(status=status_val)

        unit_status = q.get("unit_status")
        if unit_status:
            qs = qs.filter(unit_status=unit_status)

        tower_id = q.get("tower_id")
        if tower_id:
            qs = qs.filter(tower_id=tower_id)

        return qs.order_by("-id")

    # =====================================================
    # BULK CREATE (ADMIN + FULL_CONTROL SAFE)
    # =====================================================
    @action(detail=False, methods=["post"], url_path="bulk-create")
    def bulk_create(self, request):
        raw = request.data

        excel_file = request.FILES.get("file") or request.FILES.get("excel")

        if excel_file:
            try:
                df = pd.read_excel(excel_file)
            except Exception as e:
                return Response(
                    {"detail": f"Failed to read Excel: {e}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            rename_map = {
                "project_id": "project",
                "tower_id": "tower",
                "floor_id": "floor",
                "unit_id": "unit",
                "carpet_area": "carpet_sqft",
                "saleable_area": "saleable_sqft",
                "Core Base Price (psf)": "core_base_price_psf",
                "Approved Limit Price (psf)": "approved_limit_price_psf",
                "Customer Base Price (psf)": "customer_base_price_psf",
                "base_rate": "rate_psf",
                "base_amount": "agreement_value",
                "remarks": "description",
                "block_days": "block_period_days",
            }

            df.rename(columns=rename_map, inplace=True)
            df = df.where(df.notnull(), None)
            items = df.to_dict(orient="records")
            files_map = {}

        elif isinstance(raw, list):
            items = raw
            files_map = request.FILES

        else:
            items_json = raw.get("items")
            if not items_json:
                return Response(
                    {"detail": 'Expected a JSON list, "items" field, or Excel "file".'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                items = json.loads(items_json)
            except json.JSONDecodeError:
                return Response(
                    {"detail": 'Invalid JSON in "items".'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            files_map = request.FILES

        if not isinstance(items, list):
            return Response({"detail": "Items must be a list."}, status=400)

        created = []
        errors = []

        user = request.user
        role = getattr(user, "role", None)
        role_code = role.upper() if isinstance(role, str) else getattr(role, "code", None)
        admin_id = (
            user.id if role_code == "ADMIN"
            else getattr(user, "admin_id", None)
        )

        with transaction.atomic():
            for idx, payload in enumerate(items, start=1):
                payload = dict(payload)

                # 🔐 SECURITY CHECK (PROJECT OWNERSHIP)
                project_id = payload.get("project")
                if project_id and not user.is_staff:
                    if not Project.objects.filter(
                        id=project_id,
                        belongs_to_id=admin_id,
                    ).exists():
                        errors.append({
                            "index": idx,
                            "errors": {"project": ["Not allowed for this project."]}
                        })
                        continue

                docs_meta = payload.pop("documents", [])
                resolved_docs = []

                for doc in docs_meta or []:
                    file_key = doc.get("file_field")
                    file_obj = files_map.get(file_key) if file_key else None
                    if not file_obj:
                        continue
                    doc_data = dict(doc)
                    doc_data.pop("file_field", None)
                    doc_data["file"] = file_obj
                    resolved_docs.append(doc_data)

                if resolved_docs:
                    payload["documents"] = resolved_docs

                ser = InventorySerializer(data=payload, context={"request": request})
                if ser.is_valid():
                    inv = ser.save()
                    created.append(inv.id)
                else:
                    errors.append({"index": idx, "errors": ser.errors})

        code = (
            status.HTTP_201_CREATED
            if created and not errors
            else status.HTTP_207_MULTI_STATUS
        )
        return Response({"created_ids": created, "errors": errors}, status=code)

    # =====================================================
    # BY UNIT
    # =====================================================
    @action(detail=False, methods=["get"], url_path="by-unit")
    def by_unit(self, request):
        unit_id = request.query_params.get("unit_id")
        if not unit_id:
            return Response(
                {"detail": "unit_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        inv = get_object_or_404(
            Inventory.objects.select_related(
                "project",
                "tower",
                "floor",
                "unit",
                "unit_type",
                "configuration",
                "facing",
            ).prefetch_related("documents"),
            unit_id=unit_id,
        )

        serializer = InventoryDetailSerializer(
            inv, context={"request": request}
        )
        return Response(serializer.data, status=status.HTTP_200_OK)



from django.shortcuts import get_object_or_404
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status, permissions

from openpyxl import load_workbook

from .models import Project, Tower, Floor, Unit  


def normalize_status(raw):
    """
    Map Excel status text to Unit.status choices.
    Adjust mapping to match your Unit.STATUS_CHOICES.
    """
    if raw is None:
        return None
    s = str(raw).strip().upper()

    mapping = {
      "AVAILABLE": "AVAILABLE",
      "AVAIL": "AVAILABLE",
      "OPEN": "AVAILABLE",
      "BOOKED": "BOOKED",
      "BLOCKED": "BLOCKED",
      "HOLD": "BLOCKED",
      "SOLD": "SOLD",
    }
    return mapping.get(s, "AVAILABLE")  # default if unknown



class InventoryExcelUploadAPIView(APIView):
    """
    POST /api/client/inventory/upload-excel/

    Form-data:
      - project_id: <int>   (Project pk)
      - file: <Excel .xlsx>

    Each row in Excel => one Unit (inventory) under that project.
    """

    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        project_id = request.data.get("project_id")
        file_obj = request.FILES.get("file")

        if not project_id:
            return Response(
                {"detail": "project_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file_obj:
            return Response(
                {"detail": "Excel file is required (field name: 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project = get_object_or_404(Project, pk=project_id)

        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response(
                {"detail": f"Unable to read Excel file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active

        # Expected header row at row 1
        # Adjust header names if you change Excel format
        expected_headers = [
            "Tower Name",
            "Floor Number",
            "Unit Number",
            "BHK Type",
            "Carpet Area",
            "Saleable Area",
            "Status",
            "Facing",
            "Remarks",
        ]

        # Read header row
        header_row = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # (Optional) basic header check – here we only check first N columns
        for i, expected in enumerate(expected_headers):
            if i >= len(header_row):
                return Response(
                    {"detail": f"Excel is missing column '{expected}' at position {i+1}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if header_row[i].lower() != expected.lower():
                return Response(
                    {"detail": f"Invalid header in column {i+1}. Expected '{expected}', got '{header_row[i]}'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        created = 0
        updated = 0
        errors = []

        # We wrap in atomic transaction so partial import can be controlled if you want.
        # Here: we don't roll back on errors; we just log them.
        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # read cells safe
                tower_name    = (row[0].value or "").strip() if row[0].value is not None else ""
                floor_number  = (row[1].value or "").strip() if row[1].value is not None else ""
                unit_number   = (row[2].value or "").strip() if row[2].value is not None else ""
                bhk_type      = (row[3].value or "").strip() if row[3].value is not None else ""
                carpet_area   = row[4].value
                saleable_area = row[5].value
                status_raw    = row[6].value
                facing        = (row[7].value or "").strip() if len(row) > 7 and row[7].value is not None else ""
                remarks       = (row[8].value or "").strip() if len(row) > 8 and row[8].value is not None else ""

                # skip completely empty rows
                if not (tower_name or floor_number or unit_number):
                    continue

                row_errors = []

                if not tower_name:
                    row_errors.append("Tower Name is required.")
                if not floor_number:
                    row_errors.append("Floor Number is required.")
                if not unit_number:
                    row_errors.append("Unit Number is required.")

                if row_errors:
                    errors.append({"row": idx, "errors": row_errors})
                    continue

                try:
                    tower, _ = Tower.objects.get_or_create(
                        project=project,
                        name=tower_name,
                    )

                    floor, _ = Floor.objects.get_or_create(
                        project=project,
                        tower=tower,
                        number=floor_number,
                    )

                    unit_status = normalize_status(status_raw)

                    # adjust field names as per your Unit model
                    unit, created_flag = Unit.objects.update_or_create(
                        project=project,
                        tower=tower,
                        floor=floor,
                        unit_no=unit_number,
                        defaults={
                            "unit_type": bhk_type or "",
                            "carpet_area": carpet_area or 0,
                            "saleable_area": saleable_area or 0,
                            "status": unit_status,
                            "facing": facing,
                            "remarks": remarks,
                        },
                    )

                    if created_flag:
                        created += 1
                    else:
                        updated += 1

                except Exception as e:
                    errors.append({"row": idx, "errors": [str(e)]})

        return Response(
            {
                "project_id": project.id,
                "created_units": created,
                "updated_units": updated,
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )



# clientsetup/views_inventory.py

from django.shortcuts import get_object_or_404
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Project, Tower, Floor, Unit


class InventoryTreeAPIView(APIView):
    """
    GET /api/client/inventory/tree/?project_id=<id>

    Returns:
    {
      "project": {"id": ..., "name": "..."},
      "towers": [
        {
          "id": ..., "name": "...",
          "floors": [
            {
              "id": ..., "number": "...",
              "units": [
                {
                  "id": ...,
                  "unit_no": "...",
                  "inventory": {
                    "id": ...,
                    "status": "DRAFT",
                    "unit_status": "NOT_RELEASED",
                    "availability_status": "AVAILABLE",
                    "agreement_value": "12000000.00",
                    "total_cost": "13000000.00"
                  }
                }
              ]
            }
          ]
        }
      ]
    }
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        project_id = request.query_params.get("project_id")
        if not project_id:
            return Response({"detail": "project_id is required."}, status=400)

        project = get_object_or_404(Project, pk=project_id)

        # ---- Preload all structure objects for that project ----
        towers_qs = Tower.objects.filter(project_id=project.id).order_by("id")
        floors_qs = (
            Floor.objects
            .filter(tower__project_id=project.id)
            .select_related("tower")
            .order_by("id")
        )
        units_qs = (
            Unit.objects
            .filter(project_id=project.id)
            .select_related("tower", "floor")
            .order_by("id")
        )

        # ---- Preload inventory and map by unit_id ----
        inventory_qs = (
            Inventory.objects
            .filter(project_id=project.id)
            .select_related("unit")
        )
        inventory_by_unit = {inv.unit_id: inv for inv in inventory_qs}

        # ---- Build lookup dicts for floors & units ----
        floors_by_tower = {}
        for f in floors_qs:
            floors_by_tower.setdefault(f.tower_id, []).append(f)

        units_by_floor = {}
        for u in units_qs:
            units_by_floor.setdefault(u.floor_id, []).append(u)

        # ---- Build response ----
        data = {
            "project": {
                "id": project.id,
                "name": getattr(project, "name", str(project)),
            },
            "towers": [],
        }

        for t in towers_qs:
            tower_data = {
                "id": t.id,
                "name": getattr(t, "name", f"Tower {t.id}"),
                "floors": [],
            }

            for f in floors_by_tower.get(t.id, []):
                floor_data = {
                    "id": f.id,
                    "number": getattr(f, "number", getattr(f, "name", f"Floor {f.id}")),
                    "units": [],
                }

                for u in units_by_floor.get(f.id, []):
                    inv = inventory_by_unit.get(u.id)

                    unit_data = {
                        "id": u.id,
                        "unit_no": getattr(u, "unit_no", getattr(u, "name", f"Unit {u.id}")),
                        "inventory": None,
                    }

                    if inv:
                        unit_data["inventory"] = {
                            "id": inv.id,
                            "status": inv.status,
                            "unit_status": inv.unit_status,
                            "availability_status": inv.availability_status,
                            "agreement_value": inv.agreement_value,
                            "total_cost": inv.total_cost,
                            "core_base_price_psf": inv.core_base_price_psf,
                            "approved_limit_price_psf": inv.approved_limit_price_psf,
                            "customer_base_price_psf": inv.customer_base_price_psf,
                        }

                    floor_data["units"].append(unit_data)

                tower_data["floors"].append(floor_data)

            data["towers"].append(tower_data)

        return Response(data, status=200)




from rest_framework import generics, permissions
from .models import Inventory, AvailabilityStatus  # adjust import if AvailabilityStatus is in another module
from .serializers_inventory import InventoryAvailableUnitSerializer


class AvailableInventoryByProjectAPIView(generics.ListAPIView):
    """
    GET /api/client/projects/<project_id>/available-units/

    Returns inventory rows for given project where inventory is AVAILABLE.
    """

    serializer_class = InventoryAvailableUnitSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        project_id = self.kwargs.get("project_id")

        qs = (
            Inventory.objects
            .select_related(
                "project",
                "tower",
                "floor",
                "unit",
                "unit_type",
                "configuration",
                "facing",
            )
            .filter(
                project_id=project_id,
                availability_status=AvailabilityStatus.AVAILABLE,
            )
        )

        return qs



