from datetime import datetime, date
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.contrib.auth import get_user_model
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status, permissions
from openpyxl import load_workbook , Workbook
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
from accounts.models import Role
from .models import Project, Tower, Floor, ProjectStatus, ApprovalStatus, FloorStatus,Unit
from setup.models import ProjectType, TowerType
import io

User = get_user_model()


def normalize_choice(raw, choices, default):
    """Convert raw cell -> valid choice code or default."""
    if not raw:
        return default
    code = str(raw).strip().upper()
    valid_codes = {c[0] for c in choices}
    return code if code in valid_codes else default


def as_str(val: object) -> str:
    """
    Safely convert Excel cell value to a stripped string.
    Numbers, dates, etc → str()
    None → ""
    """
    if val is None:
        return ""
    return str(val).strip()



def parse_excel_date(val):
    """Convert Excel cell value to date or None."""
    if val in (None, ""):
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class ProjectExcelUploadAPIView(APIView):
    """
    POST /api/client/projects/upload-excel/

    Form-data:
      - file: .xlsx
      - admin_id: <int> (required if staff, ignored for admins)

    Each row => one Project (upsert by RERA No).
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        user = request.user

        # ---- resolve owner admin ----
        if user.is_staff:
            admin_id_raw = request.data.get("admin_id") or request.query_params.get("admin_id")
            if not admin_id_raw:
                return Response({"detail": "admin_id is required for staff import."}, status=400)
            try:
                admin_id = int(admin_id_raw)
            except ValueError:
                return Response({"detail": "admin_id must be an integer."}, status=400)
            admin = User.objects.filter(id=admin_id, role=Role.ADMIN).first()
            if not admin:
                return Response({"detail": "admin_id must refer to a valid ADMIN user."}, status=400)
        elif getattr(user, "role", None) == Role.ADMIN:
            admin = user
        else:
            return Response({"detail": "Only staff or admin can import projects."}, status=403)

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Excel file is required (field: 'file')."}, status=400)

        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response({"detail": f"Unable to read Excel file: {e}"}, status=400)

        ws = wb.active

        expected_headers = [
            "Project Name",
            "Location",
            "Developer",
            "RERA No",
            "Project Type Code",
            "Start Date",
            "End Date",
            "Possession Date",
            "Status",
            "Approval Status",
            "Notes",
        ]

        header_row = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        for i, expected in enumerate(expected_headers):
            if i >= len(header_row):
                return Response(
                    {"detail": f"Excel is missing column '{expected}' at position {i+1}."},
                    status=400,
                )
            if header_row[i].lower() != expected.lower():
                return Response(
                    {
                        "detail": f"Invalid header in column {i+1}. "
                                  f"Expected '{expected}', got '{header_row[i]}'."
                    },
                    status=400,
                )

        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                name         = (row[0].value or "").strip() if row[0].value else ""
                location     = (row[1].value or "").strip() if row[1].value else ""
                developer    = (row[2].value or "").strip() if row[2].value else ""
                rera_no      = (row[3].value or "").strip() if row[3].value else ""
                pt_code      = (row[4].value or "").strip() if row[4].value else ""
                start_raw    = row[5].value
                end_raw      = row[6].value
                poss_raw     = row[7].value
                status_raw   = row[8].value
                approval_raw = row[9].value
                notes        = (row[10].value or "").strip() if len(row) > 10 and row[10].value else ""

                if not (name or rera_no):
                    # completely empty / junk row
                    continue

                row_errors = []
                if not name:
                    row_errors.append("Project Name is required.")
                if not rera_no:
                    row_errors.append("RERA No is required.")

                if row_errors:
                    errors.append({"row": idx, "errors": row_errors})
                    continue

                # project type
                project_type = None
                if pt_code:
                    project_type = ProjectType.objects.filter(code__iexact=pt_code).first()
                    if not project_type:
                        row_errors.append(f"ProjectType with code '{pt_code}' not found.")

                start_date = parse_excel_date(start_raw)
                end_date = parse_excel_date(end_raw)
                possession_date = parse_excel_date(poss_raw)

                if start_raw and not start_date:
                    row_errors.append("Invalid Start Date format.")
                if end_raw and not end_date:
                    row_errors.append("Invalid End Date format.")
                if poss_raw and not possession_date:
                    row_errors.append("Invalid Possession Date format.")

                status_code = normalize_choice(
                    status_raw, ProjectStatus.choices, ProjectStatus.DRAFT
                )
                approval_code = normalize_choice(
                    approval_raw, ApprovalStatus.choices, ApprovalStatus.PENDING
                )

                # business date validations
                if start_date and end_date and start_date > end_date:
                    row_errors.append("Start date cannot be after End date.")
                if end_date and possession_date and end_date > possession_date:
                    row_errors.append("End date cannot be after Possession date.")

                if row_errors:
                    errors.append({"row": idx, "errors": row_errors})
                    continue

                try:
                    obj, was_created = Project.objects.update_or_create(
                        rera_no=rera_no,
                        defaults={
                            "name": name,
                            "location": location,
                            "developer": developer,
                            "project_type": project_type,
                            "start_date": start_date,
                            "end_date": end_date,
                            "possession_date": possession_date,
                            "status": status_code,
                            "approval_status": approval_code,
                            "notes": notes,
                            "belongs_to": admin,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as e:
                    errors.append({"row": idx, "errors": [str(e)]})

        return Response(
            {
                "admin_id": admin.id,
                "created_projects": created,
                "updated_projects": updated,
                "errors": errors,
            },
            status=200,
        )




class TowerExcelUploadAPIView(APIView):
    """
    POST /api/client/towers/upload-excel/

    Form-data:
      - project_id: <int>
      - file: .xlsx

    Each row => one Tower under that project (upsert by name).
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        project_id = request.data.get("project_id") or request.query_params.get("project_id")
        file_obj = request.FILES.get("file")

        if not project_id:
            return Response({"detail": "project_id is required."}, status=400)
        if not file_obj:
            return Response({"detail": "Excel file is required (field: 'file')."}, status=400)

        try:
            project_id = int(project_id)
        except ValueError:
            return Response({"detail": "project_id must be an integer."}, status=400)

        project = get_object_or_404(Project, pk=project_id)

        # permissions like your other views
        u = request.user
        if not (
            u.is_staff
            or (getattr(u, "role", None) == Role.ADMIN and project.belongs_to_id == u.id)
        ):
            return Response({"detail": "Not allowed for this project."}, status=403)

        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response({"detail": f"Unable to read Excel file: {e}"}, status=400)

        ws = wb.active

        expected_headers = [
            "Tower Name",
            "Tower Type Code",
            "Total Floors",
            "Status",
            "Notes",
        ]

        header_row = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        for i, expected in enumerate(expected_headers):
            if i >= len(header_row):
                return Response(
                    {"detail": f"Excel is missing column '{expected}' at position {i+1}."},
                    status=400,
                )
            if header_row[i].lower() != expected.lower():
                return Response(
                    {
                        "detail": f"Invalid header in column {i+1}. "
                                  f"Expected '{expected}', got '{header_row[i]}'."
                    },
                    status=400,
                )

        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                tower_name   = as_str(row[0].value)
                tower_code   = as_str(row[1].value)
                total_floors = row[2].value
                status_raw   = as_str(row[3].value)
                notes        = as_str(row[4].value) if len(row) > 4 else ""


                if not tower_name:
                    # skip completely empty rows
                    if not (tower_code or total_floors or status_raw or notes):
                        continue
                    errors.append({"row": idx, "errors": ["Tower Name is required."]})
                    continue

                row_errors = []
                ttype_obj = None
                if tower_code:
                    ttype_obj = TowerType.objects.filter(code__iexact=tower_code).first()
                    if not ttype_obj:
                        row_errors.append(f"TowerType with code '{tower_code}' not found.")

                try:
                    tf = int(total_floors) if total_floors not in (None, "") else 0
                    if tf < 0:
                        row_errors.append("Total Floors cannot be negative.")
                except (TypeError, ValueError):
                    row_errors.append("Total Floors must be an integer.")

                status_code = normalize_choice(
                    status_raw, FloorStatus.choices, FloorStatus.DRAFT
                )

                if row_errors:
                    errors.append({"row": idx, "errors": row_errors})
                    continue

                try:
                    obj, was_created = Tower.objects.update_or_create(
                        project=project,
                        name=tower_name,
                        defaults={
                            "tower_type": ttype_obj,
                            "total_floors": tf,
                            "status": status_code,
                            "notes": notes,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as e:
                    errors.append({"row": idx, "errors": [str(e)]})

        return Response(
            {
                "project_id": project.id,
                "created_towers": created,
                "updated_towers": updated,
                "errors": errors,
            },
            status=200,
        )




class FloorExcelUploadAPIView(APIView):
    """
    POST /api/client/floors/upload-excel/

    Form-data:
      - project_id: <int>
      - file: .xlsx

    Each row => one Floor under a Tower of that project.
    If tower doesn't exist, it will be created.
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        project_id = request.data.get("project_id") or request.query_params.get("project_id")
        file_obj = request.FILES.get("file")

        if not project_id:
            return Response({"detail": "project_id is required."}, status=400)
        if not file_obj:
            return Response({"detail": "Excel file is required (field: 'file')."}, status=400)

        try:
            project_id = int(project_id)
        except ValueError:
            return Response({"detail": "project_id must be an integer."}, status=400)

        project = get_object_or_404(Project, pk=project_id)

        u = request.user
        if not (
            u.is_staff
            or (getattr(u, "role", None) == Role.ADMIN and project.belongs_to_id == u.id)
        ):
            return Response({"detail": "Not allowed for this project."}, status=403)

        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response({"detail": f"Unable to read Excel file: {e}"}, status=400)

        ws = wb.active

        expected_headers = [
            "Tower Name",
            "Floor Number",
            "Total Units",
            "Status",
            "Notes",
        ]

        header_row = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        for i, expected in enumerate(expected_headers):
            if i >= len(header_row):
                return Response(
                    {"detail": f"Excel is missing column '{expected}' at position {i+1}."},
                    status=400,
                )
            if header_row[i].lower() != expected.lower():
                return Response(
                    {
                        "detail": f"Invalid header in column {i+1}. "
                                  f"Expected '{expected}', got '{header_row[i]}'."
                    },
                    status=400,
                )

        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                tower_name   = as_str(row[0].value)
                floor_number = as_str(row[1].value)
                total_units  = row[2].value  # keep as raw, we int() it later
                status_raw   = as_str(row[3].value)  # normalize_choice will get a clean string
                notes        = as_str(row[4].value) if len(row) > 4 else ""

                if not (tower_name or floor_number):
                    # empty row
                    continue

                row_errors = []
                if not tower_name:
                    row_errors.append("Tower Name is required.")
                if not floor_number:
                    row_errors.append("Floor Number is required.")

                try:
                    tu = int(total_units) if total_units not in (None, "") else 0
                    if tu < 0:
                        row_errors.append("Total Units cannot be negative.")
                except (TypeError, ValueError):
                    row_errors.append("Total Units must be an integer.")

                status_code = normalize_choice(
                    status_raw, FloorStatus.choices, FloorStatus.DRAFT
                )

                if row_errors:
                    errors.append({"row": idx, "errors": row_errors})
                    continue

                try:
                    tower, _ = Tower.objects.get_or_create(
                        project=project,
                        name=tower_name,
                        defaults={"total_floors": 0, "status": FloorStatus.DRAFT},
                    )

                    floor, was_created = Floor.objects.update_or_create(
                        tower=tower,
                        number=floor_number,
                        defaults={
                            "total_units": tu,
                            "status": status_code,
                            "notes": notes,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

                except Exception as e:
                    errors.append({"row": idx, "errors": [str(e)]})

        return Response(
            {
                "project_id": project.id,
                "created_floors": created,
                "updated_floors": updated,
                "errors": errors,
            },
            status=200,
        )



class UnitExcelUploadAPIView(APIView):
    """
    POST /api/client/units/upload-excel/
    Form-Data:
      - file: .xlsx/.xls
      - project_id: int

    Expected header row in Excel (lowercase names):
      tower_name        (required)
      floor_number      (required)
      unit_no           (required)
      unit_type_code    (optional; matches UnitType.code)
      carpet_sqft       (optional, decimal)
      builtup_sqft      (optional, decimal)
      rera_sqft         (optional, decimal)
      facing_code       (optional; matches Facing.code)
      parking_type_code (optional; matches ParkingType.code)
      agreement_value   (optional, decimal)
      construction_start (optional, yyyy-mm-dd or dd/mm/yyyy or Excel date)
      completion_date    (optional, yyyy-mm-dd or dd/mm/yyyy or Excel date)
      status_code        (optional; one of UnitStatus values, e.g. NOT_RELEASED)
      notes              (optional)
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        upload = request.FILES.get("file")
        project_id = request.data.get("project_id") or request.POST.get("project_id")

        if not upload:
            return Response({"detail": "file is required."}, status=400)
        if not project_id:
            return Response({"detail": "project_id is required."}, status=400)

        try:
            project_id = int(project_id)
        except ValueError:
            return Response({"detail": "project_id must be an integer."}, status=400)

        project = get_object_or_404(Project, id=project_id)

        # ----- permission: staff can do any; admin only own project -----
        u = request.user
        if not (
            u.is_staff
            or (getattr(u, "role", None) == Role.ADMIN and project.belongs_to_id == u.id)
        ):
            return Response({"detail": "Not allowed for this project."}, status=403)

        # ----- open workbook -----
        try:
            wb = load_workbook(upload, data_only=True)
        except Exception as exc:
            return Response({"detail": f"Unable to read workbook: {exc}"}, status=400)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=1))
        if not rows:
            return Response({"detail": "Excel file is empty."}, status=400)

        header_cells = rows[0]
        headers = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in header_cells
        ]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def get_val(row, key, default=""):
            idx = header_map.get(key)
            if idx is None:
                return default
            val = row[idx].value
            if val is None:
                return default
            if isinstance(val, str):
                return val.strip()
            return val

        def parse_decimal(value, field_name):
            if value in ("", None):
                return None
            try:
                return Decimal(str(value))
            except InvalidOperation:
                raise ValueError(f"Invalid decimal for {field_name}: {value!r}")

        def parse_date(value, field_name):
            if value in ("", None):
                return None
            if isinstance(value, (datetime.date, datetime.datetime)):
                return value.date() if isinstance(value, datetime.datetime) else value
            if isinstance(value, str):
                v = value.strip()
                # try ISO first
                try:
                    return datetime.date.fromisoformat(v)
                except ValueError:
                    # fallback dd/mm/YYYY
                    try:
                        return datetime.datetime.strptime(v, "%d/%m/%Y").date()
                    except ValueError:
                        raise ValueError(f"Invalid date for {field_name}: {value!r}")
            raise ValueError(f"Invalid date for {field_name}: {value!r}")

        created_count = 0
        updated_count = 0
        errors = []

        for idx, row in enumerate(rows[1:], start=2):
            # skip empty rows
            if all(c.value in (None, "") for c in row):
                continue

            try:
                tower_name = get_val(row, "tower_name")
                floor_number = get_val(row, "floor_number")
                unit_no = get_val(row, "unit_no")

                if not tower_name or not floor_number or not unit_no:
                    raise ValueError("tower_name, floor_number and unit_no are required.")

                # --- tower & floor lookup ----
                try:
                    tower = Tower.objects.get(project=project, name=str(tower_name))
                except Tower.DoesNotExist:
                    raise ValueError(f"Tower '{tower_name}' not found for this project.")

                try:
                    floor = Floor.objects.get(tower=tower, number=str(floor_number))
                except Floor.DoesNotExist:
                    raise ValueError(
                        f"Floor '{floor_number}' not found in tower '{tower_name}'."
                    )

                # --- get or create unit ----
                unit, created = Unit.objects.get_or_create(
                    project=project,
                    tower=tower,
                    floor=floor,
                    unit_no=str(unit_no),
                )

                # --- optional master lookups ----
                unit_type_code = get_val(row, "unit_type_code")
                if unit_type_code:
                    try:
                        ut = UnitType.objects.get(
                            code__iexact=str(unit_type_code), is_active=True
                        )
                        unit.unit_type = ut
                    except UnitType.DoesNotExist:
                        raise ValueError(f"Invalid unit_type_code '{unit_type_code}'")

                facing_code = get_val(row, "facing_code")
                if facing_code:
                    try:
                        fc = Facing.objects.get(
                            code__iexact=str(facing_code), is_active=True
                        )
                        unit.facing = fc
                    except Facing.DoesNotExist:
                        raise ValueError(f"Invalid facing_code '{facing_code}'")

                parking_code = get_val(row, "parking_type_code")
                if parking_code:
                    try:
                        pk = ParkingType.objects.get(
                            code__iexact=str(parking_code), is_active=True
                        )
                        unit.parking_type = pk
                    except ParkingType.DoesNotExist:
                        raise ValueError(
                            f"Invalid parking_type_code '{parking_code}'"
                        )

                # --- decimals ----
                unit.carpet_sqft = parse_decimal(
                    get_val(row, "carpet_sqft"), "carpet_sqft"
                )
                unit.builtup_sqft = parse_decimal(
                    get_val(row, "builtup_sqft"), "builtup_sqft"
                )
                unit.rera_sqft = parse_decimal(
                    get_val(row, "rera_sqft"), "rera_sqft"
                )
                unit.agreement_value = parse_decimal(
                    get_val(row, "agreement_value"), "agreement_value"
                )

                # --- dates ----
                unit.construction_start = parse_date(
                    get_val(row, "construction_start"), "construction_start"
                )
                unit.completion_date = parse_date(
                    get_val(row, "completion_date"), "completion_date"
                )

                # --- status ----
                status_code = get_val(row, "status_code")
                if status_code:
                    status_code = str(status_code)
                    if status_code not in UnitStatus.values:
                        raise ValueError(f"Invalid status_code '{status_code}'")
                    unit.status = status_code

                unit.notes = get_val(row, "notes", "")

                unit.save()
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as exc:
                errors.append({"row": idx, "errors": [str(exc)]})

        return Response(
            {
                "created_units": created_count,
                "updated_units": updated_count,
                "errors": errors,
            },
            status=200,
        )


def excel_response(filename, headers, sample_row=None):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    if sample_row:
        ws.append(sample_row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



class ProjectExcelSampleAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        headers = [
            "name",
            "location",
            "developer",
            "rera_no",
            "start_date",
            "end_date",
            "possession_date",
            "project_type_code",
            "status_code",
            "approval_status_code",
            "notes",
        ]
        sample_row = [
            "Star Residency",
            "Kalyan",
            "ABC Developers",
            "P1234567",
            "2025-01-01",
            "2027-12-31",
            "2028-06-30",
            "RESIDENTIAL",
            "DRAFT",
            "PENDING",
            "Phase 1 inventory",
        ]
        return excel_response("project_sample.xlsx", headers, sample_row)


class TowerExcelSampleAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        headers = [
            "project_name",
            "tower_name",
            "tower_type_code",
            "total_floors",
            "status_code",
            "notes",
        ]
        sample_row = [
            "Star Residency",
            "Tower A",
            "TOWER_HIGHRISE",
            25,
            "DRAFT",
            "Facing main road",
        ]
        return excel_response("tower_sample.xlsx", headers, sample_row)


class FloorExcelSampleAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        headers = [
            "project_name",
            "tower_name",
            "floor_number",
            "total_units",
            "status_code",
            "notes",
        ]
        sample_row = [
            "Star Residency",
            "Tower A",
            "1",
            4,
            "DRAFT",
            "Podium floor",
        ]
        return excel_response("floor_sample.xlsx", headers, sample_row)


class UnitExcelSampleAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        headers = [
            "tower_name",
            "floor_number",
            "unit_no",
            "unit_type_code",
            "carpet_sqft",
            "builtup_sqft",
            "rera_sqft",
            "facing_code",
            "parking_type_code",
            "agreement_value",
            "construction_start",
            "completion_date",
            "status_code",
            "notes",
        ]
        sample_row = [
            "Tower A",
            "1",
            "101",
            "2BHK",
            650,
            800,
            750,
            "EAST",
            "STILT",
            9500000,
            "2025-01-01",
            "2027-12-31",
            "NOT_RELEASED",
            "Corner flat",
        ]
        return excel_response("unit_sample.xlsx", headers, sample_row)


