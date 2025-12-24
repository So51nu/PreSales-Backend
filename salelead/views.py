from django.db import models as dj_models
from .timeline import build_lead_timeline,filter_timeline_events
from django.db import transaction
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import serializers as drf_serializers
from django.contrib.auth import get_user_model
from rest_framework import serializers
from django.db.models import Prefetch
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import SalesLead, LeadOpportunity, SalesLeadUpdate, SalesLeadStageHistory
# salelead/views.py (ya jaha LeadOpportunityViewSet hai)
from accounts.models import User, Role
from leadmanage.models import (
    LeadClassification,
    LeadSource,
    LeadStatus,
    LeadSubStatus,
    LeadPurpose,
    LeadStage,
)
from channel.models import (
    ChannelPartnerProfile,
    ChannelPartnerProjectAuthorization,
    ProjectAssignmentStatus,
)
from .models import SalesLeadCPInfo, SalesLeadStageHistory  # if not already

from .models import (
    SalesLead,
    SalesLeadStatusHistory,
    SalesLeadChangeLog,
    build_lead_snapshot,
    build_lead_changes,
)
import re
import pandas as pd

from django.db import transaction, models as dj_models
from django.shortcuts import get_object_or_404

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from decimal import Decimal

from clientsetup.models import Project
from channel.models import ChannelPartnerProfile
from .models import (
    SalesLead,
    SalesLeadStatusHistory,
    SalesLeadAddress,
    SalesLeadCPInfo,
)
# tumhare existing imports: LeadStatus, LeadStage, LeadOpportunity, serializers, permissions, _project_ids_for_user, _parse_bool, etc.

import os
import io
import csv
from zipfile import BadZipFile
from openpyxl import load_workbook

from django.db import transaction
from django.contrib.auth import get_user_model
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser,JSONParser

from .models import (
    LeadOpportunity,
    LeadOpportunityStatusConfig,
)
from .serializers import LeadOpportunityStatusChangeSerializer
from clientsetup.models import Project

from django.db.models import Count
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from leadmanage.models import (
    ProjectLead,
)
from salelead.models import (
    SalesLeadUpdate,
    SalesLeadUpdateStatus,
    SalesLeadUpdateStatusHistory,
)

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LeadOpportunity, SalesLead, LeadOpportunityStatus
from .serializers import LeadOpportunitySerializer
from django.db.models import Max, Count
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from .models import SiteVisit, SalesLead,SalesLeadStatusHistory
from .serializers import (
    SiteVisitSerializer,
    LeadVisitStatusUpdateSerializer,
    SiteVisitRescheduleSerializer
)
from leadmanage.models import LeadStatus,  LeadSubStatus
from rest_framework.response import Response  
from rest_framework import viewsets, permissions
from .models import LeadComment
from .serializers import LeadCommentSerializer
from .models import (
    SalesLead,
    SalesLeadUpdate,
    SalesLeadStageHistory,
    SalesLeadDocument,
    SalesLeadProposalDocument,
)
from channel.models import ChannelPartnerProfile
from django.utils import timezone
from .timeline import build_lead_timeline
from common.pagination import TimelinePagination  
from leadmanage.models import LeadStage
from salelead.models import SalesLeadStageHistory
from salelead.full_profile_serializers import SalesLeadLookupSerializer

from .serializers import (
    SalesLeadSerializer,
    SalesLeadUpdateSerializer,
    LeadStatusChangeSerializer,
    SalesLeadStageHistorySerializer,
    SalesLeadDocumentSerializer,
    SalesLeadAddressSerializer,
    SalesLeadCPInfoSerializer,
    SalesLeadPersonalInfoSerializer,
    SalesLeadProfessionalInfoSerializer,
    SalesLeadProposalDocumentSerializer,
        LeadOpportunityStatusChangeSerializer,

)
from clientsetup.models import Project,Inventory
from .full_profile_serializers import SalesLeadFullDetailSerializer
import re
from .serializers import (
    SalesLeadAddressSerializer,
    SalesLeadCPInfoSerializer,
    SalesLeadPersonalInfoSerializer,
    SalesLeadSerializer,
    SalesLeadUpdateSerializer,
    SalesLeadStageHistorySerializer,
    SalesLeadDocumentSerializer,
    SalesLeadProfessionalInfoSerializer,
    SalesLeadProposalDocumentSerializer,
)
from .models import (
    SalesLead,
    SalesLeadStatusHistory,
    SalesLeadChangeLog,
    build_lead_snapshot,
    build_lead_changes,
)
import os
import re
from uuid import uuid4
from decimal import Decimal, InvalidOperation

import pandas as pd

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.shortcuts import get_object_or_404

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

from accounts.models import User, Role
from clientsetup.models import Project
from salelead.utils import _project_ids_for_user

from channel.models import (
    ChannelPartnerProfile,
    ChannelPartnerProjectAuthorization,
    ProjectAssignmentStatus,
)

# make sure these are imported from your app where they live
from leadmanage.models import (
    LeadClassification,
    LeadSource,
    LeadStatus,
    LeadSubStatus,
    LeadPurpose,
    LeadStage,
)

from .models import (
    SalesLead,
    SalesLeadStageHistory,
    SalesLeadCPInfo,
    SalesLeadUpdate,
    SiteVisit,
)

# If you have these in another module, import them correctly.
# from channel.models import ChannelPartnerProfile
# from channel.models import ChannelPartnerProjectAuthorization, ProjectAssignmentStatus


# ✅ Add these imports at top of file (if not already)
import os
import re
from uuid import uuid4
from decimal import Decimal, InvalidOperation

import pandas as pd

from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

from accounts.models import User, Role
from clientsetup.models import Project
from salelead.utils import _project_ids_for_user

from channel.models import (
    ChannelPartnerProfile,
    ChannelPartnerProjectAuthorization,
    ProjectAssignmentStatus,
)

# make sure these imports match your project
from leadmanage.models import (
    LeadClassification,
    LeadSource,
    LeadStatus,
    LeadSubStatus,
    LeadPurpose,
    LeadStage,
)

from .models import SalesLead, SalesLeadStageHistory, SalesLeadCPInfo




def _parse_bool(val):
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    return str(val).strip().lower() in ("1", "true", "yes", "y", "on")


def _eligible_sales_users_for_project(project_id):
    """
    Yaha pe tum apna actual project-based user logic laga sakte ho.
    For now: all active SALES users.
    If you have UserAccess / project mapping, filter accordingly.
    """
    User = get_user_model()
    return User.objects.filter(role="SALES", is_active=True)


def _pick_round_robin_user(project_id):
    """
    Simple round-robin:
    - Get all eligible SALES users for this project
    - Count how many leads each has in this project
    - Choose the user with minimum count (tie-break by user id)
    """
    users_qs = _eligible_sales_users_for_project(project_id)
    if not users_qs.exists():
        raise ValidationError("No active SALES users available for round robin assignment.")

    user_ids = list(users_qs.values_list("id", flat=True))

    # Count leads per user in this project
    lead_counts_raw = (
        SalesLead.objects
        .filter(project_id=project_id, assign_to_id__in=user_ids)
        .values("assign_to_id")
        .annotate(c=Count("id"))
    )
    counts = {row["assign_to_id"]: row["c"] for row in lead_counts_raw}

    # Ensure users with 0 leads are included
    for uid in user_ids:
        counts.setdefault(uid, 0)

    # Pick user with minimum leads, then by smallest id
    chosen_user_id = sorted(counts.items(), key=lambda kv: (kv[1], kv[0]))[0][0]
    return users_qs.get(id=chosen_user_id)



class IsAuthenticatedAndActive(permissions.IsAuthenticated):
    pass


from .utils import _project_ids_for_user





from decimal import Decimal, InvalidOperation
from django.db.models import OuterRef, Subquery
from django.db.models.functions import Coalesce
class SalesLeadViewSet(viewsets.ModelViewSet):
    queryset = SalesLead.objects.select_related(
        "project", "status", "sub_status", "current_owner", "channel_partner"
    )
    serializer_class = SalesLeadSerializer
    permission_classes = [IsAuthenticatedAndActive]

    # ----------------- BASE QUERYSET + FILTERS -----------------
   
    def get_queryset(self):
        user = self.request.user
        project_ids = _project_ids_for_user(user)

        qs = (
            SalesLead.objects
            .select_related(
                "project",
                "first_owner",
                "current_owner",
                "assign_to",
                "channel_partner",
                "classification",
                "sub_classification",
                "source",
                "sub_source",
                "status",
                "sub_status",
                "purpose",
            )
            .prefetch_related(
                "offering_types",
                "updates",
                "stage_history",
                "documents",
                "project__lead_stages",
            )
            .filter(project_id__in=project_ids)   # 👈 project scope always
        )

        role = getattr(user, "role", None)
        action = getattr(self, "action", None)

        # ---------- role based lead-level filter ----------
        # ⚠️ Sirf LIST ke time restrict karna hai
        if action == "list":
            if role == "ADMIN" or user.is_staff:
                # Admin / staff => sab dekh sakte within unke projects
                pass
            elif role == "SALES":
                # SALES => sirf jaha assign_to = user (LIST me)
                qs = qs.filter(assign_to_id=user.id)
            else:
                # Others (e.g. RECEPTION) => jaha owner ya assign_to ho
                qs = qs.filter(
                    dj_models.Q(current_owner=user)
                    | dj_models.Q(assign_to=user)
                )
        # ⚠️ retrieve / detail, update, custom actions, etc. ke liye
        # yaha koi extra role filter nahi lagega, sirf project scope ka filter
        # (project_id__in=project_ids) upper already laga hua hai.

        # ---------- extra filters from query params ----------
        p = self.request.query_params

        # filter directly by project_id
        if pid := p.get("project"):
            qs = qs.filter(project_id=pid)

        if aid := p.get("assign_to"):
            qs = qs.filter(assign_to_id=aid)

        owner = p.get("owner")
        if owner == "me":
            qs = qs.filter(current_owner_id=user.id)
        elif owner and owner.isdigit():
            qs = qs.filter(current_owner_id=int(owner))

        # 🔹 filter by channel_partner
        if cp_id := p.get("channel_partner"):
            qs = qs.filter(channel_partner_id=cp_id)

        # 🔹 filter by walking flag
        walking_param = p.get("walking")
        if walking_param is not None:
            w = str(walking_param).lower()
            if w in ("1", "true", "yes", "y", "on"):
                qs = qs.filter(walking=True)
            elif w in ("0", "false", "no", "n", "off"):
                qs = qs.filter(walking=False)

        # 🔹 latest site-visit remarks per lead (for list/detail)
        latest_visit_qs = (
            SiteVisit.objects
            .filter(lead_id=OuterRef("pk"))
            .order_by("-scheduled_at", "-id")
        )

        qs = qs.annotate(
            _lv_outcome_notes=Subquery(latest_visit_qs.values("outcome_notes")[:1]),
            _lv_public_notes=Subquery(latest_visit_qs.values("public_notes")[:1]),
            _lv_internal_notes=Subquery(latest_visit_qs.values("internal_notes")[:1]),
            _lv_cancelled_reason=Subquery(latest_visit_qs.values("cancelled_reason")[:1]),
            _lv_no_show_reason=Subquery(latest_visit_qs.values("no_show_reason")[:1]),
        ).annotate(
            latest_remarks=Coalesce(
                "_lv_outcome_notes",
                "_lv_public_notes",
                "_lv_internal_notes",
                "_lv_cancelled_reason",
                "_lv_no_show_reason",
            )
        )

        return qs.order_by("-id")

    # ----------------- GENERIC CREATE LOG -----------------
    def perform_create(self, serializer):
        request = self.request
        user = request.user

        lead = serializer.save(
            created_by=user if user.is_authenticated else None
        )

        after_snapshot = build_lead_snapshot(lead)
        changes = build_lead_changes(None, after_snapshot)

        SalesLeadChangeLog.objects.create(
            sales_lead=lead,
            action=SalesLeadChangeLog.Action.CREATE,
            snapshot_before=None,
            snapshot_after=after_snapshot,
            changes=changes,
            comment=(request.data.get("comment") or "").strip()
            if isinstance(request.data, dict)
            else "",
            changed_by=user if user.is_authenticated else None,
            request_meta={
                "ip": request.META.get("REMOTE_ADDR"),
                "user_agent": request.META.get("HTTP_USER_AGENT"),
            },
        )

    # ----------------- GENERIC UPDATE LOG -----------------
    def perform_update(self, serializer):
        request = self.request
        user = request.user

        # BEFORE snapshot from current instance
        lead_before = serializer.instance
        before_snapshot = build_lead_snapshot(lead_before)

        # actual update
        lead_after = serializer.save()

        # AFTER snapshot
        after_snapshot = build_lead_snapshot(lead_after)
        changes = build_lead_changes(before_snapshot, after_snapshot)

        if not changes:
            return  # no-op update, skip log

        SalesLeadChangeLog.objects.create(
            sales_lead=lead_after,
            action=SalesLeadChangeLog.Action.UPDATE,
            snapshot_before=before_snapshot,
            snapshot_after=after_snapshot,
            changes=changes,
            comment=(request.data.get("comment") or "").strip()
            if isinstance(request.data, dict)
            else "",
            changed_by=user if user.is_authenticated else None,
            request_meta={
                "ip": request.META.get("REMOTE_ADDR"),
                "user_agent": request.META.get("HTTP_USER_AGENT"),
            },
        )

    # ----------------- BUNDLE CREATE -----------------
    @action(detail=False, methods=["post"], url_path="bundle-create")
    @transaction.atomic
    def bundle_create(self, request):
        """
        POST /api/sales/sales-leads/bundle-create/

        Payload:
        {
          "lead": {
             "project": 1,
             "first_name": "...",
             ...
             "assign_to": 5 | null,
             "round_robin": true | false   # <- write-only
          },
          "first_update": { ... }   # optional (abhi unused)
        }
        """
        lead_data = request.data.get("lead") or {}
        update_data = request.data.get("first_update") or {}

        # ---- Extract & normalise assignment inputs ----
        round_robin_raw = lead_data.pop("round_robin", False)
        round_robin = _parse_bool(round_robin_raw)

        raw_assign_to = lead_data.get("assign_to")
        has_assign_to = raw_assign_to not in (None, "", "null")

        # project is mandatory for round_robin logic
        project_id = lead_data.get("project")
        if not project_id:
            return Response(
                {"detail": "project is required for creating a lead."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ---- Validate combination: assign_to vs round_robin ----
        if not has_assign_to and not round_robin:
            return Response(
                {
                    "detail": "Either 'assign_to' must be set or 'round_robin' must be true. "
                              "At least one assignment option is required."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if has_assign_to and round_robin:
            return Response(
                {"detail": "Provide either 'assign_to' or 'round_robin', not both."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lead_ser = SalesLeadSerializer(
            data=lead_data,
            context={"request": request},
        )
        lead_ser.is_valid(raise_exception=True)
        lead = lead_ser.save(created_by=request.user)

        # --------- Round robin logic ---------
        if round_robin and not lead.assign_to_id:
            chosen_user = _pick_round_robin_user(project_id)
            lead.assign_to = chosen_user
            lead.current_owner = chosen_user
            lead.save(update_fields=["assign_to", "current_owner"])

        # --------- Walk-in -> auto stage history ---------
        if getattr(lead, "walking", False):
            site_stage = (
                LeadStage.objects.filter(
                    project_id=lead.project_id,
                    for_site=True,
                )
                .order_by("order")
                .first()
            )

            if site_stage:
                already_exists = lead.stage_history.filter(stage=site_stage).exists()
                if not already_exists:
                    from .models import SalesLeadStageHistory

                    SalesLeadStageHistory.objects.create(
                        sales_lead=lead,
                        stage=site_stage,
                        status=None,
                        sub_status=None,
                        event_date=timezone.now(),
                        created_by=request.user if request.user.is_authenticated else None,
                        notes="Auto-created for walk-in lead (for_site stage).",
                    )

        # --------- CHANGE LOG: CREATE via bundle-create ---------
        after_snapshot = build_lead_snapshot(lead)
        changes = build_lead_changes(None, after_snapshot)

        SalesLeadChangeLog.objects.create(
            sales_lead=lead,
            action=SalesLeadChangeLog.Action.CREATE,
            snapshot_before=None,
            snapshot_after=after_snapshot,
            changes=changes,
            comment="Lead created via bundle-create",
            changed_by=request.user if request.user.is_authenticated else None,
            request_meta={
                "ip": request.META.get("REMOTE_ADDR"),
                "user_agent": request.META.get("HTTP_USER_AGENT"),
            },
        )

        out_ser = SalesLeadSerializer(lead, context={"request": request})
        return Response(out_ser.data, status=status.HTTP_201_CREATED)

    # ----------------- VISIT STATUS UPDATE -----------------
    @action(detail=True, methods=["patch"], url_path="visit-status")
    def update_visit_status(self, request, pk=None):
        lead = self.get_object()
        serializer = LeadVisitStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        status_value = serializer.validated_data["status"]
        timestamp = serializer.validated_data.get("timestamp", timezone.now())

        # BEFORE snapshot
        before_snapshot = build_lead_snapshot(lead)

        lead.last_site_visit_status = status_value
        lead.last_site_visit_at = timestamp
        lead.save(update_fields=["last_site_visit_status", "last_site_visit_at"])

        # AFTER snapshot + log
        after_snapshot = build_lead_snapshot(lead)
        changes = build_lead_changes(before_snapshot, after_snapshot)

        if changes:
            SalesLeadChangeLog.objects.create(
                sales_lead=lead,
                action=SalesLeadChangeLog.Action.UPDATE,
                snapshot_before=before_snapshot,
                snapshot_after=after_snapshot,
                changes=changes,
                comment="Visit status updated via API",
                changed_by=request.user if request.user.is_authenticated else None,
                request_meta={
                    "ip": request.META.get("REMOTE_ADDR"),
                    "user_agent": request.META.get("HTTP_USER_AGENT"),
                },
            )

        return Response(
            {
                "detail": "Lead visit status updated successfully",
                "lead_id": lead.id,
                "status": status_value,
                "timestamp": timestamp,
            }
        )

    # ----------------- STATUS CHANGE -----------------
    @action(detail=True, methods=["post"], url_path="change-status")
    def change_status(self, request, pk=None):
        lead = self.get_object()

        # BEFORE
        before_snapshot = build_lead_snapshot(lead)
        old_status = getattr(lead, "status", None)
        old_sub_status = getattr(lead, "sub_status", None)

        serializer = LeadStatusChangeSerializer(
            data=request.data,
            context={"lead": lead, "request": request},
        )
        serializer.is_valid(raise_exception=True)
        lead = serializer.save()  # sets lead.status & lead.sub_status

        comment = serializer.validated_data.get("comment", "")

        # Status history row (existing audit)
        SalesLeadStatusHistory.objects.create(
            sales_lead=lead,
            old_status=old_status,
            new_status=lead.status,
            old_sub_status=old_sub_status,
            new_sub_status=lead.sub_status,
            changed_by=request.user,
            comment=comment,
        )

        # AFTER + change log
        after_snapshot = build_lead_snapshot(lead)
        changes = build_lead_changes(before_snapshot, after_snapshot)

        if changes:
            SalesLeadChangeLog.objects.create(
                sales_lead=lead,
                action=SalesLeadChangeLog.Action.STATUS_CHANGE,
                snapshot_before=before_snapshot,
                snapshot_after=after_snapshot,
                changes=changes,
                comment=comment or "Lead status changed",
                changed_by=request.user if request.user.is_authenticated else None,
                request_meta={
                    "ip": request.META.get("REMOTE_ADDR"),
                    "user_agent": request.META.get("HTTP_USER_AGENT"),
                },
            )

        return Response(
            {
                "detail": "Lead status updated successfully.",
                "lead_id": lead.id,
                "project_id": lead.project_id,
                "status": lead.status.name if lead.status else None,
                "sub_status": lead.sub_status.name if lead.sub_status else None,
            },
            status=status.HTTP_200_OK,
        )

    # ----------------- RETRIEVE + PROJECT STATUSES -----------------
    def retrieve(self, request, *args, **kwargs):
        """
        GET /api/sales/sales-leads/<id>/?include_all_stage=true

        If include_all_stage=true, also return all LeadStatus for this lead's project.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data

        include_all_stage = request.query_params.get("include_all_stage")
        flag = str(include_all_stage).lower() if include_all_stage is not None else ""

        if flag in ("1", "true", "yes", "y", "on"):
            # All statuses for THIS lead's project
            statuses_qs = LeadStatus.objects.filter(
                project_id=instance.project_id
            ).order_by("name")

            data["project_statuses"] = [
                {
                    "id": s.id,
                    "name": s.name,
                }
                for s in statuses_qs
            ]

        return Response(data)

    # ----------------- FULL INFO -----------------
    @action(detail=True, methods=["get"], url_path="full-info")
    def full_info(self, request, pk=None):
        qs = self.filter_queryset(
            self.get_queryset()
            .select_related(
                "project",
                "first_owner",
                "current_owner",
                "assign_to",
                "channel_partner",
                "classification",
                "sub_classification",
                "source",
                "sub_source",
                "status",
                "sub_status",
                "purpose",
                "address",
                "cp_info",
                "personal_info",
                "professional_info",
            )
            .prefetch_related(
                "offering_types",
                "proposal_documents",
                "interested_unit_links__unit",
                Prefetch(
                    "payments",
                    queryset=PaymentLead.objects.filter(for_kyc=False).order_by(
                        "-payment_date", "-id"
                    ),
                ),
            )
        )

        lead = get_object_or_404(qs, pk=pk)
        serializer = SalesLeadFullDetailSerializer(lead, context={"request": request})
        return Response(serializer.data)

    # ----------------- LOOKUP BY PHONE -----------------
    @action(detail=False, methods=["get"], url_path="lookup-by-phone")
    def lookup_by_phone(self, request):
        """
        GET /api/sales/sales-leads/lookup-by-phone/?phone=9876543210&project_id=3

        Behaviour:
          - uses project_id to find its owner (project.belongs_to)
          - collects all projects with same owner the user has scope on
          - searches leads + opportunities by mobile across those projects
          - DOES NOT apply role-based queryset filters (created_by/assign_to/etc)
          - returns counts + first 5 leads/opps
        """
        raw_phone = (request.query_params.get("phone") or "").strip()
        project_id_raw = request.query_params.get("project_id")

        # basic validation
        if not raw_phone:
            return Response(
                {"detail": "phone is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        digits = re.sub(r"\D", "", raw_phone)
        if len(digits) != 10:
            return Response(
                {"detail": "phone must be a 10-digit number."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            project_id = int(project_id_raw) if project_id_raw else None
        except ValueError:
            return Response(
                {"detail": "project_id must be integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # base scope: projects this user is allowed to see
        user_project_ids = set(_project_ids_for_user(request.user))

        if not user_project_ids:
            return Response(
                {
                    "present": False,
                    "lead_count": 0,
                    "opportunity_count": 0,
                    "leads": [],
                    "opportunities": [],
                },
                status=status.HTTP_200_OK,
            )

        # ---- derive project_ids from owner (belongs_to) ----
        if project_id:
            if project_id not in user_project_ids:
                # user not allowed on this project
                return Response(
                    {
                        "present": False,
                        "lead_count": 0,
                        "opportunity_count": 0,
                        "leads": [],
                        "opportunities": [],
                    },
                    status=status.HTTP_200_OK,
                )

            try:
                base_project = Project.objects.get(pk=project_id)
            except Project.DoesNotExist:
                return Response(
                    {"detail": "Invalid project_id."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            belongs_to_id = getattr(base_project, "belongs_to_id", None)

            if belongs_to_id:
                owner_project_ids = set(
                    Project.objects.filter(
                        belongs_to_id=belongs_to_id
                    ).values_list("id", flat=True)
                )
                project_ids = list(owner_project_ids & user_project_ids)
                if not project_ids:
                    project_ids = [project_id]
            else:
                project_ids = [project_id]
        else:
            # no project_id provided → use all scoped projects
            project_ids = list(user_project_ids)

        # --------- Leads (NO role-based filter here) ----------
        lead_qs = (
            SalesLead.objects
            .select_related("project", "status", "sub_status", "current_owner", "channel_partner")
            .filter(
                project_id__in=project_ids,
                mobile_number__icontains=digits,
            )
            .order_by("-created_at")
        )

        lead_data = SalesLeadLookupSerializer(
            lead_qs[:5], many=True, context={"request": request}
        ).data

        # --------- Opportunities ----------
        opp_qs = (
            LeadOpportunity.objects.filter(
                project_id__in=project_ids,
                mobile_number__icontains=digits,
            )
            .order_by("-created_at")
        )

        opp_data = LeadOpportunitySerializer(
            opp_qs[:5], many=True, context={"request": request}
        ).data

        present = lead_qs.exists() or opp_qs.exists()

        return Response(
            {
                "present": present,
                "lead_count": lead_qs.count(),
                "opportunity_count": opp_qs.count(),
                "leads": lead_data,
                "opportunities": opp_data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["post"], url_path="copy-missing")
    @transaction.atomic
    def copy_missing(self, request):
        """
        POST /api/sales/sales-leads/copy-missing/

        Body:
          {
            "from_lead_id": <old lead id>,
            "to_lead_id": <new lead id>
          }

        Behaviour:
          - For Address + ProfessionalInfo:
              * if target section doesn't exist → clone from source
              * else copy only fields that are empty on target & filled on source
        """
        from_id = request.data.get("from_lead_id")
        to_id = request.data.get("to_lead_id")

        if not from_id or not to_id:
            return Response(
                {"detail": "from_lead_id and to_lead_id are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if str(from_id) == str(to_id):
            return Response(
                {"detail": "from_lead_id and to_lead_id must be different."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # scope check
        allowed_project_ids = set(_project_ids_for_user(request.user))

        try:
            from_lead = SalesLead.objects.get(pk=from_id)
            to_lead = SalesLead.objects.get(pk=to_id)
        except SalesLead.DoesNotExist:
            return Response(
                {"detail": "One or both leads not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if (
            from_lead.project_id not in allowed_project_ids
            or to_lead.project_id not in allowed_project_ids
        ):
            return Response(
                {"detail": "You are not allowed to modify one of these leads."},
                status=status.HTTP_403_FORBIDDEN,
            )

        result = {
            "from_lead_id": from_lead.id,
            "to_lead_id": to_lead.id,
            "address": {"created": False, "copied_fields": []},
            "professional_info": {"created": False, "copied_fields": []},
        }

        # ---------- Address ----------
        source_addr = getattr(from_lead, "address", None)
        if source_addr:
            try:
                target_addr = to_lead.address
                addr_created = False
            except SalesLeadAddress.DoesNotExist:
                target_addr = SalesLeadAddress(sales_lead=to_lead)
                addr_created = True

            addr_fields = [
                "flat_or_building",
                "area",
                "pincode",
                "city",
                "state",
                "country",
                "description",
            ]
            addr_copied = []

            for field in addr_fields:
                src_val = getattr(source_addr, field)
                tgt_val = getattr(target_addr, field)
                if (not tgt_val) and src_val:
                    setattr(target_addr, field, src_val)
                    addr_copied.append(field)

            if addr_created or addr_copied:
                target_addr.save()

            result["address"]["created"] = addr_created
            result["address"]["copied_fields"] = addr_copied

        # ---------- Professional Info ----------
        source_prof = getattr(from_lead, "professional_info", None)
        if source_prof:
            try:
                target_prof = to_lead.professional_info
                prof_created = False
            except SalesLeadProfessionalInfo.DoesNotExist:
                target_prof = SalesLeadProfessionalInfo(sales_lead=to_lead)
                prof_created = True

            prof_fields = [
                "occupation",
                "organization_name",
                "office_location",
                "office_pincode",
                "designation",
            ]
            prof_copied = []

            for field in prof_fields:
                src_val = getattr(source_prof, field)
                tgt_val = getattr(target_prof, field)
                # FKs (occupation/designation) can be None vs actual object
                if (not tgt_val) and src_val:
                    setattr(target_prof, field, src_val)
                    prof_copied.append(field)

            if prof_created or prof_copied:
                target_prof.save()

            result["professional_info"]["created"] = prof_created
            result["professional_info"]["copied_fields"] = prof_copied

        return Response(result, status=status.HTTP_200_OK)

    # ----------------- TIMELINE -----------------
    @action(detail=True, methods=["get"], url_path="timeline")
    def timeline(self, request, pk=None):
        """
        GET /api/sales/sales-leads/<id>/timeline/

        Filters:
          - ?types=BOOKING,PAYMENT,SITE_VISIT,LEAD_CHANGE,...
          - ?from_date=2025-11-01
          - ?to_date=2025-11-30
          - ?q=booking

        Pagination:
          - ?page=1
          - ?page_size=25
        """
        lead = self.get_object()

        # 1) sab events build karo
        events = build_lead_timeline(lead)

        # 2) filters apply
        events = filter_timeline_events(events, request)

        # 3) paginate
        paginator = TimelinePagination()
        page = paginator.paginate_queryset(events, request, view=self)
        return paginator.get_paginated_response(page)

    # =====================================================================
    #                     SECTION CRUD (Address / CP / Personal / Prof / Docs)

    @action(detail=True, methods=["get", "put", "patch"], url_path="address")
    def address(self, request, pk=None):
        """
        GET  /sales-leads/<id>/address/
        PUT  /sales-leads/<id>/address/
        PATCH/same
        """
        lead = self.get_object()
        try:
            instance = lead.address
        except SalesLeadAddress.DoesNotExist:
            instance = None

        if request.method == "GET":
            if not instance:
                return Response(
                    {
                        "id": None,
                        "flat_or_building": "",
                        "area": "",
                        "pincode": "",
                        "city": "",
                        "state": "",
                        "country": "",
                        "description": "",
                    }
                )
            ser = SalesLeadAddressSerializer(instance)
            return Response(ser.data)

        partial = request.method == "PATCH"
        ser = SalesLeadAddressSerializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        obj = ser.save(sales_lead=lead)
        return Response(SalesLeadAddressSerializer(obj).data)



    # ✅ FULL FUNCTION
    @action(detail=False, methods=["post"], url_path="import-excel")
    @transaction.atomic
    def import_excel(self, request):
        user = request.user

        # ---- Project check (URL or body) ----
        project_id_raw = request.query_params.get("project_id") or request.data.get("project_id")
        if not project_id_raw:
            return Response(
                {"detail": "project_id is required in URL (?project_id=...) or body."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            project_id = int(project_id_raw)
        except ValueError:
            return Response({"detail": "project_id must be integer."}, status=status.HTTP_400_BAD_REQUEST)

        # Scope check: user is allowed on this project?
        allowed_project_ids = _project_ids_for_user(user)
        if project_id not in allowed_project_ids:
            return Response(
                {"detail": "You are not allowed to import leads for this project."},
                status=status.HTTP_403_FORBIDDEN,
            )

        project = get_object_or_404(Project, pk=project_id)

        # ---- Excel/CSV file ----
        excel_file = request.FILES.get("file") or request.FILES.get("excel")
        if not excel_file:
            return Response(
                {"detail": 'Excel/CSV file is required (use field name "file").'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Read into DataFrame: support .xlsx / .xls / .csv
        try:
            name = getattr(excel_file, "name", "") or ""
            content_type = getattr(excel_file, "content_type", "") or ""
            ext = os.path.splitext(name)[1].lower()

            if ext == ".csv" or "csv" in content_type.lower():
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file, engine="openpyxl")
        except Exception as e:
            return Response({"detail": f"Failed to read Excel/CSV: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if df.empty:
            return Response({"detail": "Excel/CSV file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        # ------------------------------------------------------------------
        # ✅ HEADER NORMALIZATION (THIS FIXES YOUR CP NOT CREATING / NOT MAPPING)
        # Your excel headers like "CP Mobile Number" were becoming "cp mobile number"
        # and your code was looking for "cp_mobile_number".
        # ------------------------------------------------------------------
        def norm_header(h):
            h = str(h).strip().lower()
            h = re.sub(r"[^a-z0-9]+", "_", h).strip("_")  # spaces/dots -> _
            return h

        df.columns = [norm_header(c) for c in df.columns]

        # ✅ Map your exact Excel columns -> code fields
        # Your headers:
        # Client Name, Number, Email Address, Company Name, Walk-in Source, Sub Source, Sub source 1,
        # CP Firm Name, CP Name, CP Mobile Number, CP Rera No., CP Firm Address
        df.rename(
            columns={
                "number": "mobile_number",
                "email_address": "email",
                "company_name": "company",
                "walk_in_source": "source",
                "sub_source": "sub_source",
                "sub_source_1": "sub_source_1",

                "cp_firm_name": "cp_company",
                "cp_name": "cp_name",
                "cp_mobile_number": "cp_mobile_number",
                "cp_rera_no": "cp_rera_number",
                "cp_firm_address": "cp_address",
            },
            inplace=True,
        )

        # ✅ If file doesn't have first_name/last_name, derive from client_name
        if "first_name" not in df.columns:
            if "client_name" in df.columns:
                names = df["client_name"].fillna("").astype(str).str.strip()
                df["first_name"] = names.str.split().str[0].fillna("")
                df["last_name"] = names.apply(lambda x: " ".join(x.split()[1:]) if x else "")
            else:
                df["first_name"] = ""
                df["last_name"] = ""

        # ---- Required columns (now guaranteed) ----
        required_cols = ["first_name", "mobile_number"]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # NaN -> None
        df = df.where(df.notnull(), None)
        rows = df.to_dict(orient="records")

        # ---------- Helpers ----------
        def norm_mobile(val):
            if not val:
                return ""
            return re.sub(r"\D", "", str(val))

        def s(val):
            if val is None:
                return ""
            text = str(val).strip()
            if text.lower() in {"nan", "none", "null"}:
                return ""
            return text

        def resolve_admin_owner_for_cp():
            # Prefer logged-in ADMIN
            if getattr(user, "role", None) == Role.ADMIN:
                return user
            # else user's admin
            admin_obj = getattr(user, "admin", None)
            if admin_obj and getattr(admin_obj, "role", None) == Role.ADMIN:
                return admin_obj
            # else project.admin / project.created_by (if exists)
            proj_admin = getattr(project, "admin", None)
            if proj_admin and getattr(proj_admin, "role", None) == Role.ADMIN:
                return proj_admin
            proj_created_by = getattr(project, "created_by", None)
            if proj_created_by and getattr(proj_created_by, "role", None) == Role.ADMIN:
                return proj_created_by
            return None

        admin_owner = resolve_admin_owner_for_cp()

        # ---- Existing leads in this project (duplicate check) ----
        existing_pairs = set()  # (email_lower, mobile_norm)
        for r in SalesLead.objects.filter(project_id=project_id).values("email", "mobile_number"):
            e = (r["email"] or "").strip().lower()
            m = norm_mobile(r["mobile_number"])
            if e or m:
                existing_pairs.add((e, m))

        # ---- Preload taxonomies for this project ----
        class_qs = LeadClassification.objects.filter(project=project)
        classification_map = {c.name.strip().lower(): c for c in class_qs.filter(parent__isnull=True)}
        sub_class_map = {(c.parent_id, c.name.strip().lower()): c for c in class_qs.filter(parent__isnull=False)}

        source_qs = LeadSource.objects.filter(project=project)
        source_map = {x.name.strip().lower(): x for x in source_qs.filter(parent__isnull=True)}
        sub_source_map = {(x.parent_id, x.name.strip().lower()): x for x in source_qs.filter(parent__isnull=False)}

        status_qs = LeadStatus.objects.filter(project=project)
        status_map = {x.name.strip().lower(): x for x in status_qs}

        sub_status_qs = LeadSubStatus.objects.filter(status__project=project)
        sub_status_map = {(x.status_id, x.name.strip().lower()): x for x in sub_status_qs}

        purpose_qs = LeadPurpose.objects.filter(project=project)
        purpose_map = {x.name.strip().lower(): x for x in purpose_qs}

        stage_qs = LeadStage.objects.filter(project=project)
        stage_map = {x.name.strip().lower(): x for x in stage_qs}

        # ---- Preload assigned-to users via email ----
        assign_to_email_set = set()
        for r in rows:
            assign_email_str = s(
                r.get("assign_to_email")
                or r.get("assigned_to_email")
                or r.get("owner_email")
                or r.get("assign_to")
            )
            if assign_email_str:
                assign_to_email_set.add(assign_email_str.lower())

        assign_users_by_email = {}
        if assign_to_email_set:
            qs = User.objects.filter(email__in=list(assign_to_email_set))
            assign_users_by_email = {u.email.lower(): u for u in qs}

        created_ids = []
        errors = []

        for idx, r in enumerate(rows, start=2):
            channel_partner = None

            first_name = s(r.get("first_name"))
            last_name = s(r.get("last_name"))

            email_raw = s(r.get("email"))
            email = email_raw or None

            mobile_raw = r.get("mobile_number") or r.get("mobile") or r.get("phone") or r.get("phone_number")
            mobile = norm_mobile(mobile_raw)

            row_errors = []
            row_warnings = []

            if not first_name and not last_name:
                row_errors.append("first_name or last_name is required.")
            if not mobile:
                row_errors.append("mobile_number is required.")

            if row_errors:
                errors.append({"row": idx, "name": f"{first_name} {last_name}".strip() or None, "errors": row_errors})
                continue

            key = ((email or "").lower(), mobile)
            if key in existing_pairs:
                errors.append(
                    {
                        "row": idx,
                        "name": f"{first_name} {last_name}".strip() or None,
                        "errors": ["Lead with same email & mobile already exists in this project."],
                    }
                )
                continue

            company = s(r.get("company"))[:150]

            # ---- budget ----
            raw_budget = r.get("budget")
            budget = None
            if raw_budget not in (None, ""):
                txt = str(raw_budget).strip()
                if txt.lower() not in {"nan", "none", "null"}:
                    txt = txt.replace(",", "")
                    try:
                        budget = Decimal(txt)
                    except InvalidOperation:
                        row_warnings.append(f'Budget "{txt}" is not a valid decimal; treating as blank.')

            # ---- annual_income ----
            raw_income = r.get("annual_income")
            annual_income = None
            if raw_income not in (None, ""):
                txt = str(raw_income).strip()
                if txt.lower() not in {"nan", "none", "null"}:
                    txt = txt.replace(",", "")
                    try:
                        annual_income = Decimal(txt)
                    except InvalidOperation:
                        row_warnings.append(f'Annual income "{txt}" is not a valid decimal; treating as blank.')

            walking_raw = r.get("walking")
            walking = False
            if walking_raw is not None:
                w_str = s(walking_raw).lower()
                if w_str in ("true", "false"):
                    walking = (w_str == "true")

            tel_res = s(r.get("tel_res") or r.get("tel_residence"))
            tel_office = s(r.get("tel_office") or r.get("telephone_office"))

            # ---- taxonomy lookups ----
            classification_obj = None
            sub_classification_obj = None
            classification_name = s(r.get("classification"))
            if classification_name:
                classification_obj = classification_map.get(classification_name.lower())

            sub_class_name = s(r.get("sub_classification"))
            if sub_class_name and classification_obj:
                sub_classification_obj = sub_class_map.get((classification_obj.id, sub_class_name.lower()))

            source_obj = None
            sub_source_obj = None
            source_name = s(r.get("source"))
            if source_name:
                source_obj = source_map.get(source_name.lower())

            # ✅ your file may have sub_source_1
            sub_source_name = s(r.get("sub_source") or r.get("sub_source_1"))
            if sub_source_name and source_obj:
                sub_source_obj = sub_source_map.get((source_obj.id, sub_source_name.lower()))

            status_obj = None
            sub_status_obj = None
            status_name = s(r.get("status"))
            if status_name:
                status_obj = status_map.get(status_name.lower())

            sub_status_name = s(r.get("sub_status"))
            if sub_status_name and status_obj:
                sub_status_obj = sub_status_map.get((status_obj.id, sub_status_name.lower()))

            purpose_obj = None
            purpose_name = s(r.get("purpose"))
            if purpose_name:
                purpose_obj = purpose_map.get(purpose_name.lower())

            stage_obj = None
            stage_name = s(r.get("stage") or r.get("lead_stage"))
            if stage_name:
                stage_obj = stage_map.get(stage_name.lower())

            # ------------------------------------------------------------------
            # ✅ CP RESOLVE / CREATE / AUTHORIZE (using your actual excel headers)
            # ------------------------------------------------------------------
            cp_company = s(r.get("cp_company"))
            cp_name = s(r.get("cp_name"))
            cp_mobile = norm_mobile(r.get("cp_mobile_number"))
            cp_rera = s(r.get("cp_rera_number"))
            cp_address = s(r.get("cp_address"))
            cp_email = s(r.get("cp_email") or r.get("channel_partner_email")).lower()

            # 1) Find existing profile by mobile/company
            q = Q()
            if cp_mobile:
                q |= Q(mobile_number=cp_mobile)
            if cp_company:
                q |= Q(company_name__iexact=cp_company)

            if q:
                channel_partner = ChannelPartnerProfile.objects.filter(q).select_related("user").first()

            # 2) fallback by email -> CP user -> profile
            if not channel_partner and cp_email:
                cp_user_existing = User.objects.filter(email__iexact=cp_email, role=Role.CP).first()
                if cp_user_existing:
                    channel_partner = ChannelPartnerProfile.objects.filter(user=cp_user_existing).select_related("user").first()

            # 3) create if not found AND cp info exists
            if not channel_partner and (cp_mobile or cp_company or cp_email):
                # email required in your User model
                cp_email_final = cp_email or f"cp_{uuid4().hex}@placeholder.local"

                # if user exists with same email, reuse
                cp_user = User.objects.filter(email__iexact=cp_email_final).first()
                if not cp_user:
                    username_final = f"cp_{cp_mobile or idx}_{uuid4().hex[:6]}"
                    cp_user = User.objects.create(
                        username=username_final,
                        email=cp_email_final,
                        first_name=cp_name or "",
                        role=Role.CP,
                        is_active=True,
                        admin=admin_owner,  # IMPORTANT for your system
                        created_by=user if user.is_authenticated else None,
                    )
                    cp_user.set_unusable_password()
                    cp_user.save(update_fields=["password"])

                # create profile
                channel_partner = ChannelPartnerProfile.objects.create(
                    user=cp_user,
                    mobile_number=cp_mobile or "",
                    company_name=cp_company or "",
                    address=cp_address or "",
                    rera_number=cp_rera or "",
                    created_by=user if user.is_authenticated else None,
                    last_modified_by=user if user.is_authenticated else None,
                    last_modified_at=timezone.now(),
                )

            # 4) authorize to project
            if channel_partner:
                ChannelPartnerProjectAuthorization.objects.get_or_create(
                    channel_partner=channel_partner,
                    project=project,
                    defaults={
                        "status": ProjectAssignmentStatus.ACTIVE,
                        "created_by": user if user.is_authenticated else None,
                    },
                )

            # ---- assigned to (optional) ----
            assign_to_user = None
            assign_email_str = s(
                r.get("assign_to_email")
                or r.get("assigned_to_email")
                or r.get("owner_email")
                or r.get("assign_to")
            )
            if assign_email_str:
                assign_to_user = assign_users_by_email.get(assign_email_str.lower())
                if not assign_to_user:
                    row_warnings.append(f"Assigned user '{assign_email_str}' not found; assigning to current user.")
                    assign_to_user = user if user.is_authenticated else None
            else:
                assign_to_user = user if user.is_authenticated else None

            # ---- Create lead ----
            lead = SalesLead.objects.create(
                project=project,
                first_name=first_name,
                last_name=last_name,
                email=email,
                mobile_number=mobile,
                company=company,
                budget=budget,
                annual_income=annual_income,
                walking=walking,
                unknown_channel_partner=None,
                tel_res=tel_res or "",
                tel_office=tel_office or "",
                classification=classification_obj,
                sub_classification=sub_classification_obj,
                source=source_obj,
                sub_source=sub_source_obj,
                status=status_obj,
                sub_status=sub_status_obj,
                purpose=purpose_obj,
                channel_partner=channel_partner,  # ✅ mapped
                assign_to=assign_to_user,
                current_owner=assign_to_user if assign_to_user else None,
                first_owner=assign_to_user if assign_to_user else None,
                created_by=user if user.is_authenticated else None,
            )

            # per-lead cp info
            if channel_partner:
                SalesLeadCPInfo.objects.update_or_create(
                    sales_lead=lead,
                    defaults={"cp_user": channel_partner.user},
                )

            created_ids.append(lead.id)
            existing_pairs.add(key)

            # ---- stage history ----
            if stage_obj:
                already_exists = lead.stage_history.filter(stage=stage_obj).exists()
                if not already_exists:
                    SalesLeadStageHistory.objects.create(
                        sales_lead=lead,
                        stage=stage_obj,
                        status=status_obj,
                        sub_status=sub_status_obj,
                        event_date=timezone.now(),
                        created_by=user if user.is_authenticated else None,
                        notes="Stage set from Excel import.",
                    )

            if row_warnings:
                errors.append({"row": idx, "name": f"{first_name} {last_name}".strip() or None, "errors": row_warnings})

        status_code = status.HTTP_201_CREATED if created_ids and not errors else status.HTTP_207_MULTI_STATUS
        return Response(
            {
                "project_id": project.id,
                "created_count": len(created_ids),
                "created_ids": created_ids,
                "error_count": len(errors),
                "errors": errors,
            },
            status=status_code,
        )
        # ---------- CP INFO (1–1) ----------
   
   
    @action(detail=True, methods=["get", "put", "patch"], url_path="cp-info")
    def cp_info(self, request, pk=None):
        lead = self.get_object()
        try:
            instance = lead.cp_info
        except SalesLeadCPInfo.DoesNotExist:
            instance = None

        if request.method == "GET":
            if not instance:
                return Response({"id": None, "cp_user": None, "referral_code": ""})
            ser = SalesLeadCPInfoSerializer(instance)
            return Response(ser.data)

        partial = request.method == "PATCH"
        ser = SalesLeadCPInfoSerializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        obj = ser.save(sales_lead=lead)
        return Response(SalesLeadCPInfoSerializer(obj).data)

    # ---------- PERSONAL INFO (1–1) ----------
    @action(detail=True, methods=["get", "put", "patch"], url_path="personal-info")
    def personal_info(self, request, pk=None):
        lead = self.get_object()
        try:
            instance = lead.personal_info
        except SalesLeadPersonalInfo.DoesNotExist:
            instance = None

        if request.method == "GET":
            if not instance:
                return Response(
                    {
                        "id": None,
                        "date_of_birth": None,
                        "date_of_anniversary": None,
                        "already_part_of_family": False,
                        "secondary_email": "",
                        "alternate_mobile": "",
                        "alternate_tel_res": "",
                        "alternate_tel_off": "",
                        "visiting_on_behalf": None,
                        "current_residence_ownership": None,
                        "current_residence_type": "",
                        "family_size": None,
                        "possession_desired_in": None,
                        "facebook": "",
                        "twitter": "",
                        "linkedin": "",
                    }
                )
            ser = SalesLeadPersonalInfoSerializer(instance)
            return Response(ser.data)

        partial = request.method == "PATCH"
        ser = SalesLeadPersonalInfoSerializer(
            instance, data=request.data, partial=partial
        )
        ser.is_valid(raise_exception=True)
        obj = ser.save(sales_lead=lead)
        return Response(SalesLeadPersonalInfoSerializer(obj).data)

    # ---------- PROFESSIONAL INFO (1–1) ----------
    @action(detail=True, methods=["get", "put", "patch"], url_path="professional-info")
    def professional_info(self, request, pk=None):
        lead = self.get_object()
        try:
            instance = lead.professional_info
        except SalesLeadProfessionalInfo.DoesNotExist:
            instance = None

        if request.method == "GET":
            if not instance:
                return Response(
                    {
                        "id": None,
                        "occupation": None,
                        "organization_name": "",
                        "office_location": "",
                        "office_pincode": "",
                        "designation": None,
                    }
                )
            ser = SalesLeadProfessionalInfoSerializer(instance)
            return Response(ser.data)

        partial = request.method == "PATCH"
        ser = SalesLeadProfessionalInfoSerializer(
            instance, data=request.data, partial=partial
        )
        ser.is_valid(raise_exception=True)
        obj = ser.save(sales_lead=lead)
        return Response(SalesLeadProfessionalInfoSerializer(obj).data)

    # ---------- PROPOSAL DOCUMENTS (multi) ----------
    @action(detail=True, methods=["get", "post"], url_path="proposal-docs")
    def proposal_docs(self, request, pk=None):
        """
        GET  /sales-leads/<id>/proposal-docs/    -> list
        POST /sales-leads/<id>/proposal-docs/    -> upload {file}
        """
        lead = self.get_object()

        if request.method == "GET":
            qs = lead.proposal_documents.all().order_by("-created_at")
            ser = SalesLeadProposalDocumentSerializer(qs, many=True)
            return Response(ser.data)

        # POST (upload)
        ser = SalesLeadProposalDocumentSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        obj = ser.save(sales_lead=lead)
        return Response(
            SalesLeadProposalDocumentSerializer(obj).data,
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=["delete"],
        url_path=r"proposal-docs/(?P<doc_id>\d+)",
    )
    def delete_proposal_doc(self, request, pk=None, doc_id=None):
        """
        DELETE /sales-leads/<id>/proposal-docs/<doc_id>/
        """
        lead = self.get_object()
        try:
            doc = lead.proposal_documents.get(pk=doc_id)
        except SalesLeadProposalDocument.DoesNotExist:
            return Response(
                {"detail": "Document not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        doc.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)







class SalesLeadUpdateViewSet(viewsets.ModelViewSet):
    serializer_class = SalesLeadUpdateSerializer
    permission_classes = [IsAuthenticatedAndActive]

    def get_queryset(self):
        user = self.request.user
        project_ids = _project_ids_for_user(user)

        qs = (
            SalesLeadUpdate.objects
            .select_related(
                "sales_lead",
                "sales_lead__project",
                "created_by",
                "activity_status",   # 👈
            )
            .filter(sales_lead__project_id__in=project_ids)
        )

        lead_id = self.request.query_params.get("sales_lead")
        if lead_id:
            qs = qs.filter(sales_lead_id=lead_id)

        type_filter = self.request.query_params.get("update_type")
        if type_filter:
            qs = qs.filter(update_type=type_filter)

        status_id = self.request.query_params.get("activity_status")
        if status_id:
            qs = qs.filter(activity_status_id=status_id)

        return qs.order_by("-event_date", "-id")

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


    @action(detail=True, methods=["post"], url_path="change-status")
    def change_status(self, request, pk=None):
        """
        POST /api/sales/sales-lead-updates/<id>/change-status/

        Payload:
        {
          "activity_status": <status_id>,   # required
          "comment": "optional remark"      # optional
        }
        """
        update_obj = self.get_object()
        old_status = update_obj.activity_status

        class _StatusChangeSerializer(serializers.Serializer):
            activity_status = serializers.PrimaryKeyRelatedField(
                queryset=SalesLeadUpdateStatus.objects.all(),
                required=True,
            )
            comment = serializers.CharField(
                required=False,
                allow_blank=True,
                allow_null=True,
            )

        ser = _StatusChangeSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        new_status = ser.validated_data["activity_status"]
        comment = ser.validated_data.get("comment", "")

        lead_project_id = update_obj.sales_lead.project_id
        if new_status.project_id != lead_project_id:
            return Response(
                {
                    "detail": "Selected activity_status does not belong to this lead's project."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        update_obj.activity_status = new_status
        update_obj.save(update_fields=["activity_status", "updated_at"])

        SalesLeadUpdateStatusHistory.objects.create(
            sales_lead_update=update_obj,
            old_status=old_status,
            new_status=new_status,
            changed_by=request.user,
            comment=comment or "",
        )

        out_ser = self.get_serializer(update_obj)
        return Response(out_ser.data, status=status.HTTP_200_OK)



class SalesLeadStageHistoryViewSet(viewsets.ModelViewSet):
    serializer_class = SalesLeadStageHistorySerializer
    permission_classes = [IsAuthenticatedAndActive]

    def get_queryset(self):
        user = self.request.user
        project_ids = _project_ids_for_user(user)

        qs = (
            SalesLeadStageHistory.objects
            .select_related(
                "sales_lead",
                "sales_lead__project",
                "stage",
                "status",
                "sub_status",
                "created_by",
            )
            .filter(sales_lead__project_id__in=project_ids)
        )

        if lead_id := self.request.query_params.get("sales_lead"):
            qs = qs.filter(sales_lead_id=lead_id)

        return qs.order_by("-event_date", "-id")

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)



class SalesLeadDocumentViewSet(viewsets.ModelViewSet):
    """
    /api/sales/sales-lead-documents/      [GET, POST]
    /api/sales/sales-lead-documents/{id}/ [GET, PATCH, DELETE]

    Optional filter:
      ?sales_lead=<lead_id>
    """
    serializer_class = SalesLeadDocumentSerializer
    permission_classes = [IsAuthenticatedAndActive]

    def get_queryset(self):
        user = self.request.user
        project_ids = _project_ids_for_user(user)

        qs = (
            SalesLeadDocument.objects
            .select_related(
                "sales_lead",
                "sales_lead__project",
            )
            .filter(sales_lead__project_id__in=project_ids)
        )

        # filter by lead if needed
        lead_id = self.request.query_params.get("sales_lead")
        if lead_id:
            qs = qs.filter(sales_lead_id=lead_id)

        return qs.order_by("-created_at", "-id")

    def perform_create(self, serializer):
        sales_lead_id = self.request.data.get("sales_lead")
        if not sales_lead_id:
            raise drf_serializers.ValidationError(
                {"sales_lead": ["This field is required."]}
            )
        return serializer.save(sales_lead_id=sales_lead_id)


class SalesLeadExtraInfoBulkAPIView(APIView):
    """
    POST /api/salelead/sales-leads/extra-info/

    Body (JSON):
    {
      "sales_lead_id": 10,

      "address": {
        "flat_or_building": "...",
        "area": "...",
        "pincode": "...",
        "city": "...",
        "state": "...",
        "country": "...",
        "description": "..."
      },

      "cp_info": {
        "cp_user": 5,
        "referral_code": "ABC123"
      },

      "personal_info": {
        "date_of_birth": "1995-01-01",
        "date_of_anniversary": "2020-01-01",
        "already_part_of_family": true,
        "secondary_email": "...",
        "alternate_mobile": "...",
        "alternate_tel_res": "...",
        "alternate_tel_off": "...",
        "visiting_on_behalf": 1,
        "current_residence_ownership": 2,
        "current_residence_type": "Owned",
        "family_size": 3,
        "possession_desired_in": 1,
        "facebook": "",
        "twitter": "",
        "linkedin": ""
      },

      "professional_info": {
        "occupation": 1,
        "organization_name": "ABC Ltd",
        "office_location": "Mumbai",
        "office_pincode": "400001",
        "designation": 2
      }
    }

    Files (optional, multipart):
      proposal_files[] = <file1>, <file2>, ...
    """

    permission_classes = [IsAuthenticatedAndActive]

    def get(self, request, *args, **kwargs):
        """
        GET /api/sales/extra-info/?project_id=2

        Returns all additional-info master options for this project.
        """
        project_id = request.query_params.get("project_id")
        if not project_id:
            return Response(
                {"detail": "project_id query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            pl = ProjectLead.objects.select_related("project").get(
                project_id=project_id
            )
        except ProjectLead.DoesNotExist:
            # Project exists but no lead-setup yet → return empty lists
            return Response(
                {
                    "project": {"id": int(project_id), "name": None},
                    "visiting_half": [],
                    "family_size": [],
                    "residency_ownership": [],
                    "possession_designed": [],
                    "occupations": [],
                    "designations": [],
                },
                status=status.HTTP_200_OK,
            )

        def simple(qs):
            # NamedLookup usually has id + name
            return list(qs.values("id", "name"))

        data = {
            "project": {
                "id": pl.project_id,
                "name": pl.project.name,
            },
            "visiting_half": simple(pl.visiting_half_options.all()),
            "family_size": simple(pl.family_size_options.all()),
            "residency_ownership": simple(pl.residency_ownership_options.all()),
            "possession_designed": simple(pl.possession_designed_options.all()),
            "occupations": simple(pl.occupation_options.all()),
            "designations": simple(pl.designation_options.all()),
        }
        return Response(data, status=status.HTTP_200_OK)

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.data
        sales_lead_id = data.get("sales_lead_id") or data.get("sales_lead")

        if not sales_lead_id:
            return Response(
                {"sales_lead_id": ["This field is required."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lead = get_object_or_404(SalesLead, pk=sales_lead_id)

        response_payload = {"sales_lead": lead.id}
        errors = {}

        # ---------- Address ----------
        address_data = data.get("address")
        if address_data is not None and address_data != "":
            # If coming as JSON string in multipart
            if isinstance(address_data, str):
                try:
                    import json

                    address_data = json.loads(address_data)
                except Exception:
                    address_data = {}

            addr_instance = getattr(lead, "address", None)
            addr_serializer = SalesLeadAddressSerializer(
                instance=addr_instance,
                data=address_data,
                partial=True,
            )
            if addr_serializer.is_valid():
                addr_obj = addr_serializer.save(sales_lead=lead)
                response_payload["address"] = SalesLeadAddressSerializer(
                    addr_obj
                ).data
            else:
                errors["address"] = addr_serializer.errors

        # ---------- CP Info ----------
        cp_info_data = data.get("cp_info")
        if cp_info_data is not None and cp_info_data != "":
            if isinstance(cp_info_data, str):
                try:
                    import json

                    cp_info_data = json.loads(cp_info_data)
                except Exception:
                    cp_info_data = {}

            cp_instance = getattr(lead, "cp_info", None)
            cp_serializer = SalesLeadCPInfoSerializer(
                instance=cp_instance,
                data=cp_info_data,
                partial=True,
            )
            if cp_serializer.is_valid():
                cp_obj = cp_serializer.save(sales_lead=lead)
                response_payload["cp_info"] = SalesLeadCPInfoSerializer(
                    cp_obj
                ).data
            else:
                errors["cp_info"] = cp_serializer.errors

        # ---------- Personal Info (Additional Information) ----------
        personal_data = data.get("personal_info")
        if personal_data is not None and personal_data != "":
            if isinstance(personal_data, str):
                try:
                    import json

                    personal_data = json.loads(personal_data)
                except Exception:
                    personal_data = {}

            p_instance = getattr(lead, "personal_info", None)
            p_serializer = SalesLeadPersonalInfoSerializer(
                instance=p_instance,
                data=personal_data,
                partial=True,
            )
            if p_serializer.is_valid():
                p_obj = p_serializer.save(sales_lead=lead)
                response_payload["personal_info"] = (
                    SalesLeadPersonalInfoSerializer(p_obj).data
                )
            else:
                errors["personal_info"] = p_serializer.errors

        # ---------- Professional Info ----------
        professional_data = data.get("professional_info")
        if professional_data is not None and professional_data != "":
            if isinstance(professional_data, str):
                try:
                    import json

                    professional_data = json.loads(professional_data)
                except Exception:
                    professional_data = {}

            pr_instance = getattr(lead, "professional_info", None)
            pr_serializer = SalesLeadProfessionalInfoSerializer(
                instance=pr_instance,
                data=professional_data,
                partial=True,
            )
            if pr_serializer.is_valid():
                pr_obj = pr_serializer.save(sales_lead=lead)
                response_payload["professional_info"] = (
                    SalesLeadProfessionalInfoSerializer(pr_obj).data
                )
            else:
                errors["professional_info"] = pr_serializer.errors

        # ---------- Proposal Files (optional, multiple) ----------
        files = request.FILES.getlist("proposal_files")
        if files:
            created_docs = []
            for f in files:
                doc = SalesLeadProposalDocument.objects.create(
                    sales_lead=lead, file=f
                )
                created_docs.append(doc)

            response_payload["proposal_documents"] = (
                SalesLeadProposalDocumentSerializer(
                    created_docs, many=True, context={"request": request}
                ).data
            )

        # ---------- Errors? ----------
        if errors:
            return Response(
                {"sales_lead_id": lead.id, "errors": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(response_payload, status=status.HTTP_200_OK)


class LeadCommentViewSet(viewsets.ModelViewSet):
    """
    /sales/lead-comments/
      GET ?sales_lead=<id>         -> comments for one lead
      GET ?stage=<stage_id>        -> optionally filter by stage
      POST { sales_lead, text, stage_at_time } -> create
    """
    serializer_class = LeadCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = LeadComment.objects.select_related(
            "sales_lead", "created_by", "stage_at_time"
        )
        params = self.request.query_params

        lead_id = params.get("sales_lead")
        if lead_id:
            qs = qs.filter(sales_lead_id=lead_id)

        stage_id = params.get("stage")
        if stage_id:
            qs = qs.filter(stage_at_time_id=stage_id)

        return qs.order_by("-created_at")

from django.db.models import Q, Count, Max
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import SiteVisit, SiteVisitRescheduleHistory
from .serializers import (
    SiteVisitSerializer,
    LeadVisitStatusUpdateSerializer,
    SiteVisitRescheduleSerializer,
    OnsiteRegistrationSerializer,
)
from salelead.models import SalesLeadUpdate



class SiteVisitViewSet(viewsets.ModelViewSet):
    """
    Full PRE-Sale level SiteVisit viewset:
    - role-based access
    - status update (COMPLETED / CANCELLED / NO_SHOW / SCHEDULED)
    - reschedule + reschedule history
    - summary per-lead
    """

    queryset = SiteVisit.objects.select_related(
        "lead", "project", "unit_config", "inventory",
        "created_by", "updated_by"
    ).order_by("-scheduled_at")

    serializer_class = SiteVisitSerializer
    permission_classes = [IsAuthenticatedAndActive]


    # 🔹 COMMON helper – har jagah same pattern ka note
    def _log_site_visit_update(self, visit, title, base_info, extra_note=None, event_date=None):
        from salelead.models import SalesLeadUpdate  # ya top pe import rakho

        info = base_info
        note = (extra_note or "").strip()
        if note:
            info = f"{base_info} Note: {note}"

        SalesLeadUpdate.objects.create(
            sales_lead=visit.lead,
            update_type="SITE_VISIT",
            title=title,
            info=info,
            created_by=self.request.user if self.request.user.is_authenticated else None,
            event_date=event_date or timezone.now(),
        )
    # ------------------ ACCESS + FILTERS ------------------

    def get_queryset(self):
        qs = self.queryset
        request = self.request
        user = request.user

        # Project scope (ADMIN / SALES dono ke liye)
        project_ids = _project_ids_for_user(user)
        if project_ids:
            qs = qs.filter(project_id__in=project_ids)

        role = (getattr(user, "role", "") or "").upper()

        # Role based filter:
        # SALES => sirf self-created site visits
        # ADMIN / SUPER_ADMIN / staff => apne scope ke saare
        if role == "SALES":
            qs = qs.filter(created_by=user)
        elif role in ("ADMIN", "SUPER_ADMIN") or user.is_staff:
            # full within their projects
            pass
        else:
            # default: show only created_by user
            qs = qs.filter(created_by=user)

        params = request.query_params

        # search on lead name / created_by
        search = params.get("search")
        if search:
            qs = qs.filter(
                Q(lead__first_name__icontains=search)
                | Q(lead__last_name__icontains=search)
                | Q(created_by__first_name__icontains=search)
                | Q(created_by__last_name__icontains=search)
            )

        # status filter
        status_val = params.get("status")
        if status_val:
            qs = qs.filter(status=status_val)

        # project filter (within allowed)
        project = params.get("project")
        if project:
            qs = qs.filter(project_id=project)

        # created_by filter (for admin)
        created_by = params.get("created_by")
        if created_by:
            qs = qs.filter(created_by_id=created_by)

        # lead filter (for "view all visits of this lead")
        lead = params.get("lead")
        if lead:
            qs = qs.filter(lead_id=lead)

        # date range filter (scheduled date)
        start = params.get("start_date")
        end = params.get("end_date")
        if start:
            qs = qs.filter(scheduled_at__date__gte=start)
        if end:
            qs = qs.filter(scheduled_at__date__lte=end)

        return qs

    # ------------------ CREATE / UPDATE HOOKS ------------------

    def perform_create(self, serializer):
        visit = serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user,
        )

        # 🔹 Lead summary update
        lead = visit.lead
        lead.last_site_visit_status = visit.status           # "SCHEDULED"
        lead.last_site_visit_at = visit.scheduled_at
        lead.save(update_fields=["last_site_visit_status", "last_site_visit_at"])

        # 🔹 Note on timeline
        self._log_site_visit_update(
            visit,
            title="Site visit scheduled",
            base_info=(
                f"Site visit #{visit.id} scheduled on "
                f"{visit.scheduled_at.strftime('%d-%m-%Y %H:%M')} "
                f"for project {visit.project.name}."
            ),
        )


    def perform_update(self, serializer):
        visit = serializer.save(updated_by=self.request.user)

        # If someone edits scheduled_at manually, still keep latest on lead
        lead = visit.lead
        if lead:
            lead.last_site_visit_status = visit.status
            lead.last_site_visit_at = visit.scheduled_at
            lead.save(update_fields=["last_site_visit_status", "last_site_visit_at"])

    # ------------------ BY-LEAD LIST ------------------

    # GET /site-visits/by-lead/<lead_id>/
    @action(detail=False, methods=["get"], url_path="by-lead/(?P<lead_id>[^/.]+)")
    def by_lead(self, request, lead_id=None):
        qs = (
            self.get_queryset()
            .filter(lead_id=lead_id)
            .order_by("-scheduled_at", "-id")
        )
        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)
    # ------------------ STATUS UPDATE (ONE VISIT) ------------------


# PATCH /site-visits/<id>/update-status/
    @action(detail=True, methods=["patch"], url_path="update-status")
    def update_status(self, request, pk=None):
        visit = self.get_object()
        ser = LeadVisitStatusUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        status_value = ser.validated_data["status"]
        timestamp = ser.validated_data.get("timestamp", timezone.now())

        # note ya cancelled_reason me jo bhi aaye usko use karo
        note = (
            ser.validated_data.get("note", "") 
            or ser.validated_data.get("cancelled_reason", "") 
            or ""
        ).strip()

        # 🔹 status ke hisaab se REASON fields update
        if status_value == "CANCELLED":
            # cancel hone par cancelled_reason set, no_show_reason clear
            if hasattr(visit, "cancelled_reason"):
                visit.cancelled_reason = note
            if hasattr(visit, "no_show_reason"):
                visit.no_show_reason = ""
        elif status_value == "NO_SHOW":
            # no show par no_show_reason set, cancelled_reason clear
            if hasattr(visit, "no_show_reason"):
                visit.no_show_reason = note
            if hasattr(visit, "cancelled_reason"):
                visit.cancelled_reason = ""
        else:
            # baaki status me reasons optionally clear
            if hasattr(visit, "cancelled_reason"):
                visit.cancelled_reason = ""
            if hasattr(visit, "no_show_reason"):
                visit.no_show_reason = ""

        # 🔹 NOTE ko hamesha internal_notes me bhi store karo
        if note and hasattr(visit, "internal_notes"):
            # append pattern: purana + newline + [datetime] note
            if visit.internal_notes:
                visit.internal_notes = (
                    f"{visit.internal_notes}\n\n"
                    f"[{timestamp.strftime('%d-%m-%Y %H:%M')}] {note}"
                )
            else:
                visit.internal_notes = note

        visit.status = status_value
        visit.updated_by = request.user

        # 🔹 sirf EXISTING fields hi update_fields me daalo
        update_fields = ["status", "updated_by"]
        if hasattr(visit, "cancelled_reason"):
            update_fields.append("cancelled_reason")
        if hasattr(visit, "no_show_reason"):
            update_fields.append("no_show_reason")
        if hasattr(visit, "internal_notes"):
            update_fields.append("internal_notes")

        visit.save(update_fields=update_fields)

        # 🔹 Lead summary
        lead = visit.lead
        lead.last_site_visit_status = status_value
        lead.last_site_visit_at = timestamp
        lead.save(update_fields=["last_site_visit_status", "last_site_visit_at"])

        # 🔹 Title + base message decide karo
        if status_value == "COMPLETED":
            title = "Site visit completed"
            base_info = (
                f"Site visit #{visit.id} marked as COMPLETED on "
                f"{timestamp.strftime('%d-%m-%Y %H:%M')}."
            )
        elif status_value == "CANCELLED":
            title = "Site visit cancelled"
            base_info = (
                f"Site visit #{visit.id} CANCELLED on "
                f"{timestamp.strftime('%d-%m-%Y %H:%M')}."
            )
        elif status_value == "NO_SHOW":
            title = "Site visit marked as no show"
            base_info = (
                f"Site visit #{visit.id} marked as NO_SHOW on "
                f"{timestamp.strftime('%d-%m-%Y %H:%M')}."
            )
        elif status_value in ("SCHEDULED", "RESCHEDULED"):
            title = "Site visit status updated"
            base_info = (
                f"Site visit #{visit.id} status set to {status_value} on "
                f"{timestamp.strftime('%d-%m-%Y %H:%M')}."
            )
        else:
            title = "Site visit status updated"
            base_info = f"Site visit #{visit.id} status updated to {status_value}."

        # 🔹 Hamesha ek note entry timeline pe
        self._log_site_visit_update(
            visit,
            title=title,
            base_info=base_info,
            extra_note=note,
            event_date=timestamp,
        )

        return Response(SiteVisitSerializer(visit).data)
    # ------------------ SUMMARY (PER LEAD) ------------------

    # ------------------ SUMMARY (PER LEAD) ------------------


    # GET /site-visits/summary/
    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        """
        Summary per lead:
        - latest visit info (per lead)
        - total visits count
        - latest remarks
        Respect current user's role, filters, projects.
        """

        base_qs = self.get_queryset()

        agg = (
            base_qs.values("lead_id")
            .annotate(
                latest_scheduled_at=Max("scheduled_at"),
                total_visits=Count("id"),
            )
            .order_by("-latest_scheduled_at")
        )

        result = []
        for row in agg:
            lead_id = row["lead_id"]

            latest_visit = (
                base_qs.filter(lead_id=lead_id)
                .select_related("lead", "project")
                .order_by("-scheduled_at", "-id")
                .first()
            )
            if not latest_visit:
                continue

            lead_obj = latest_visit.lead
            if lead_obj:
                first = lead_obj.first_name or ""
                last = lead_obj.last_name or ""
                lead_name = (
                    f"{first} {last}".strip()
                    or getattr(lead_obj, "full_name", "")
                )
                mobile = lead_obj.mobile_number
            else:
                lead_name = ""
                mobile = None

            # 👇 pick the latest remarks / notes
            latest_remarks = (
                latest_visit.outcome_notes
                or latest_visit.public_notes
                or latest_visit.internal_notes
                or latest_visit.cancelled_reason
                or latest_visit.no_show_reason
            )

            result.append(
                {
                    "lead_id": lead_id,
                    "lead_name": lead_name,
                    "mobile": mobile,
                    "project": latest_visit.project.name if latest_visit.project else None,
                    "latest_scheduled_at": latest_visit.scheduled_at,
                    "latest_status": latest_visit.status,
                    "total_visits": row["total_visits"],
                    "latest_remarks": latest_remarks,  # 👈 NEW
                }
            )

        return Response(result)    # ------------------ RESCHEDULE (ONE VISIT) ------------------

    # POST /site-visits/<id>/reschedule/
    @action(detail=True, methods=["post"], url_path="reschedule")
    def reschedule(self, request, pk=None):
        """
        POST /api/sales/site-visits/<id>/reschedule/

        {
          "new_scheduled_at": "2025-12-01T16:00:00+05:30",
          "reason": "Customer requested next day"
        }
        """
        visit = self.get_object()

        # COMPLETED visit ko reschedule nahi karna
        if visit.status == "COMPLETED":
            return Response(
                {"detail": "Cannot reschedule a COMPLETED visit. Please create a new site visit."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = SiteVisitRescheduleSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        new_time = ser.validated_data["new_scheduled_at"]
        reason = ser.validated_data.get("reason", "").strip()
        old_time = visit.scheduled_at

        # 🔹 History table me pure trail
        SiteVisitRescheduleHistory.objects.create(
            site_visit=visit,
            old_scheduled_at=old_time,
            new_scheduled_at=new_time,
            reason=reason,
            created_by=request.user if request.user.is_authenticated else None,
        )

        # 🔹 Main visit update: status = RESCHEDULED
        visit.scheduled_at = new_time
        visit.status = "RESCHEDULED"      # 👈 yeh main change hai
        visit.updated_by = request.user
        visit.save(update_fields=["scheduled_at", "status", "updated_by"])

        # 🔹 Lead summary
        lead = visit.lead
        lead.last_site_visit_status = "RESCHEDULED"
        lead.last_site_visit_at = new_time
        lead.save(update_fields=["last_site_visit_status", "last_site_visit_at"])

        # 🔹 Timeline note
        base_info = (
            f"Site visit #{visit.id} rescheduled from "
            f"{old_time.strftime('%d-%m-%Y %H:%M')} "
            f"to {new_time.strftime('%d-%m-%Y %H:%M')}."
        )

        self._log_site_visit_update(
            visit,
            title="Site visit rescheduled",
            base_info=base_info,
            extra_note=reason,            # yahi tumhara note / reason
            event_date=timezone.now(),
        )

        return Response(SiteVisitSerializer(visit).data, status=status.HTTP_200_OK)

    # ------------------ RESCHEDULE HISTORY (FULL LIST) ------------------

    # GET /site-visits/<id>/reschedule-history/
    @action(detail=True, methods=["get"], url_path="reschedule-history")
    def reschedule_history(self, request, pk=None):
        visit = self.get_object()
        history_qs = (
            visit.reschedule_history
            .select_related("created_by")
            .order_by("-created_at")
        )

        items = []
        for h in history_qs:
            user = h.created_by
            created_by_name = (
                user.get_full_name() or user.username
                if user else None
            )
            items.append(
                {
                    "id": h.id,
                    "old_scheduled_at": h.old_scheduled_at,
                    "new_scheduled_at": h.new_scheduled_at,
                    "reason": h.reason,
                    "created_at": h.created_at,
                    "created_by_name": created_by_name,
                }
            )

        return Response(
            {
                "visit_id": visit.id,
                "count": len(items),
                "history": items,
            }
        )




from django.db.models import Q, Count 
from rest_framework.pagination import PageNumberPagination 
from datetime import datetime, timedelta
from django.utils import timezone
from openpyxl import load_workbook
from .models import LeadOpportunityAttachment
from django.db import transaction
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import viewsets, permissions
from .models import LeadOpportunityStatusConfig
from .serializers import LeadOpportunityStatusConfigSerializer
from django.db.models import Q
import csv
import io
    
UserModel = get_user_model()
from django.db import transaction
from django.contrib.auth import get_user_model

from rest_framework.decorators import action
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework import status



class LeadOpportunityPagination(PageNumberPagination):
    page_size = 10                      
    page_size_query_param = "page_size" 
    max_page_size = 100










class LeadOpportunityViewSet(viewsets.ModelViewSet):
    serializer_class = LeadOpportunitySerializer
    pagination_class = LeadOpportunityPagination
    permission_classes = [permissions.IsAuthenticated]


    def get_queryset(self):
        qs = (
            LeadOpportunity.objects
            .select_related("project", "created_by", "status_config")
            .order_by("-created_at")
        )

        request = self.request
        qp = request.query_params

        # ---------- Multi-project filter ----------
        project_params = qp.getlist("project")
        project_ids = []
        for val in project_params:
            for part in val.split(","):
                part = part.strip()
                if part:
                    project_ids.append(part)
        if project_ids:
            qs = qs.filter(project_id__in=project_ids)

        # ---------- (optional) raw status filter ----------
        status_val = qp.get("status")
        if status_val:
            qs = qs.filter(status=status_val)

        # ---------- Status-config filters ----------
        status_cfg_id = qp.get("status_config_id")
        if status_cfg_id:
            qs = qs.filter(status_config_id=status_cfg_id)

        status_cfg_code = qp.get("status_config_code")
        if status_cfg_code:
            qs = qs.filter(status_config__code=status_cfg_code)

        # ---------- Source system filter ----------
        source_system = qp.get("source_system")
        if source_system:
            qs = qs.filter(source_system=source_system)

        # ---------- Search filter ----------
        search = qp.get("search")
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search)
                | Q(email__icontains=search)
                | Q(mobile_number__icontains=search)
            )

        # ---------- DATE RANGE filter ----------
        date_from = qp.get("date_from")
        date_to = qp.get("date_to")

        if date_from:
            try:
                dt_from = datetime.strptime(date_from, "%Y-%m-%d")
                dt_from = timezone.make_aware(dt_from, timezone.get_current_timezone())
                qs = qs.filter(created_at__gte=dt_from)
            except ValueError:
                pass

        if date_to:
            try:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
                dt_to = timezone.make_aware(dt_to, timezone.get_current_timezone())
                qs = qs.filter(created_at__lt=dt_to)
            except ValueError:
                pass

        return qs
    

    @action(detail=True, methods=["post"], url_path="change-status")
    def change_status(self, request, pk=None):
        """
        POST /api/sales/lead-opportunities/{id}/change-status/

        Body:
        {
          "status_config_id": 12,
          "comment": "optional reason"
        }
        """
        opp = self.get_object()

        ser = LeadOpportunityStatusChangeSerializer(
            data=request.data,
            context={
                "request": request,
                "opportunity": opp,
            },
        )
        ser.is_valid(raise_exception=True)
        result = ser.save()

        status_cfg = result["status_config"]
        sales_lead = result["sales_lead"]

        return Response(
            {
                "detail": "Status updated successfully.",
                "id": opp.id,
                "status_config_id": status_cfg.id if status_cfg else None,
                "status_config_code": status_cfg.code if status_cfg else None,
                "status_config_label": status_cfg.label if status_cfg else None,
                "status_can_convert": bool(status_cfg.can_convert) if status_cfg else False,
                "auto_converted": result["auto_converted"],
                "sales_lead_id": sales_lead.id if sales_lead else None,
            },
            status=status.HTTP_200_OK,
        )


    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()

        # ------- Summary BEFORE pagination (full filtered queryset) -------
        total = qs.count()
        status_rows = qs.values("status").annotate(count=Count("id"))
        by_status = {row["status"]: row["count"] for row in status_rows}

        # ------- Apply pagination -------
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data["summary"] = {
                "total": total,
                "by_status": by_status,
            }
            return response

        serializer = self.get_serializer(qs, many=True)
        return Response(
            {
                "results": serializer.data,
                "summary": {
                    "total": total,
                    "by_status": by_status,
                },
            }
        )


    @action(detail=True, methods=["post"])
    def convert(self, request, pk=None):
        """
        POST /api/sales/lead-opportunities/{id}/convert/

        Optional body:
        {
          "assign_to_id": 123   # optional
        }
        """
        opp = self.get_object()

        # already converted?
        try:
            existing_lead = opp.sales_lead
        except SalesLead.DoesNotExist:
            existing_lead = None

        if existing_lead:
            return Response(
                {
                    "detail": "Already converted.",
                    "sales_lead_id": existing_lead.id,
                },
                status=status.HTTP_200_OK,
            )

        if not opp.project:
            return Response(
                {"detail": "Project is required on opportunity before conversion."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- yahan se owner logic add kiya ---
        request_user = (
            request.user
            if getattr(request, "user", None) and request.user.is_authenticated
            else None
        )

        assign_to = None
        assign_to_id = request.data.get("assign_to_id")

        if assign_to_id:
            # 1️⃣ UI se explicit assign_to_id aaya -> highest priority
            try:
                assign_to = User.objects.get(pk=assign_to_id)
            except User.DoesNotExist:
                return Response(
                    {"detail": "Invalid assign_to_id."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif opp.owner_id:
            # 2️⃣ Agar opportunity ke paas owner hai -> use that
            assign_to = opp.owner
        else:
            # 3️⃣ Fallback: current logged-in user
            assign_to = request_user

        # --- SalesLead create ---
        lead = SalesLead.objects.create(
            first_name=opp.full_name or "",
            email=opp.email or "",
            mobile_number=opp.mobile_number or "",
            project=opp.project,
            source_opportunity=opp,
            created_by=request_user,   # ya yahan chaho to opp.created_by bhi use kar sakte ho
            assign_to=assign_to,
            # Agar tumhare SalesLead model me yeh fields hain
            # to uncomment karke owner flow aur strong ban sakta hai:
            # current_owner=assign_to,
            # first_owner=request_user or assign_to,
        )

        opp.status = LeadOpportunityStatus.CONVERTED
        opp.save(update_fields=["status"])

        return Response(
            {"sales_lead_id": lead.id},
            status=status.HTTP_201_CREATED,
        )

    def perform_create(self, serializer):
        """
        For manual created opportunities:
        - auto-fill created_by
        - auto-generate a UNIQUE external_id if it's blank
        """

        request = self.request
        data = serializer.validated_data

        # use source_system from data or default
        source_system = data.get("source_system") or LeadSourceSystem.CALLING

        # Project (coming from project_id write-only field, if you added that)
        project = data.get("project")

        # Mobile number for uniqueness
        mobile = (data.get("mobile_number") or "").strip()

        # If serializer ever allows external_id in future, respect it
        external_id = (data.get("external_id") or "").strip() if "external_id" in data else ""

        if not external_id:
            # build a base external_id similar style to your import logic
            safe_mobile = "".join(ch for ch in mobile if ch.isdigit()) or "NA"
            project_part = project.id if project else "NA"

            base_external_id = f"MANUAL-{source_system}-{project_part}-{safe_mobile}"
            external_id = base_external_id
            suffix = 1

            # ensure uniqueness for (source_system, external_id)
            while LeadOpportunity.objects.filter(
                source_system=source_system,
                external_id=external_id,
            ).exists():
                external_id = f"{base_external_id}#{suffix}"
                suffix += 1

        serializer.save(
            created_by=request.user if request.user.is_authenticated else None,
            source_system=source_system,
            external_id=external_id,
        )



    @action(
        detail=False,
        methods=["post"],
        url_path="import-opportunities",
        parser_classes=[JSONParser, FormParser, MultiPartParser],
    )
    def import_opportunities(self, request, *args, **kwargs):
        """
        POST /api/sales/lead-opportunities/import-opportunities/

        multipart/form-data:
          - file: Excel (.xlsx) ya CSV

        Expected columns (header row):
          project_id, source_system, source_name, external_id,
          full_name, email, mobile_number,
          to_lead, owner_username, owner_email, remark, status_code
        """

        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "file is required (Excel/CSV with opportunities)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ---------- Decide Excel vs CSV ----------
        ext = os.path.splitext(upload.name)[1].lower()

        def _csv_reader_from_bytes(raw_bytes):
            try:
                decoded = raw_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                decoded = raw_bytes.decode("latin-1")
            io_obj = io.StringIO(decoded)
            return csv.DictReader(io_obj)

        # final "reader" = iterable of dict rows
        reader = None

        if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            # Try as real Excel
            try:
                wb = load_workbook(upload, data_only=True)
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return Response(
                        {"detail": "Uploaded Excel file is empty."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                header_row = rows[0]
                headers = [str(c or "").strip() for c in header_row]
                data_rows = rows[1:]

                def excel_dict_rows():
                    for r in data_rows:
                        row_dict = {}
                        for idx, col_name in enumerate(headers):
                            if not col_name:
                                continue
                            val = r[idx] if idx < len(r) else ""
                            row_dict[col_name] = "" if val is None else str(val)
                        yield row_dict

                reader = excel_dict_rows()

            except BadZipFile:
                # Extension .xlsx hai but file actually CSV / text hai
                upload.seek(0)
                raw_bytes = upload.read()
                reader = _csv_reader_from_bytes(raw_bytes)
        else:
            # Treat as CSV / text
            raw_bytes = upload.read()
            reader = _csv_reader_from_bytes(raw_bytes)

        truthy_values = {"true", "yes", "y", "1", "qualified", "hot", "interested"}

        processed = []
        errors = []
        skipped = []

        for idx, row in enumerate(reader, start=2):  # 2 => first data row
            row_no = idx

            # --- basic fields (strip + safe) ---
            project_id_raw = (row.get("project_id") or "").strip()
            full_name = (row.get("full_name") or "").strip()
            email = (row.get("email") or "").strip()
            mobile = (row.get("mobile_number") or "").strip()

            to_lead_raw = (row.get("to_lead") or "").strip().lower()
            owner_username = (row.get("owner_username") or "").strip()
            owner_email = (row.get("owner_email") or "").strip()
            remark = (row.get("remark") or "").strip()
            status_code_raw = (row.get("status_code") or "").strip().upper()

            source_system_raw = (row.get("source_system") or "").strip().upper()
            external_id_raw = (row.get("external_id") or "").strip()
            source_name_raw = (row.get("source_name") or "").strip()

            # ---- REQUIRED FIELDS VALIDATION ----
            if not project_id_raw:
                errors.append(
                    {"row": row_no, "reason": "project_id is required", "data": row}
                )
                continue

            if not full_name:
                errors.append(
                    {"row": row_no, "reason": "full_name is required", "data": row}
                )
                continue

            # if not email:
            #     errors.append(
            #         {"row": row_no, "reason": "email is required", "data": row}
            #     )
            #     continue

            if not mobile:
                errors.append(
                    {"row": row_no, "reason": "mobile_number is required", "data": row}
                )
                continue

            # ---- project resolve ----
            try:
                project_id = int(project_id_raw)
            except ValueError:
                errors.append(
                    {
                        "row": row_no,
                        "reason": f'Invalid project_id="{project_id_raw}"',
                        "data": row,
                    }
                )
                continue

            try:
                project = Project.objects.get(pk=project_id)
            except Project.DoesNotExist:
                errors.append(
                    {
                        "row": row_no,
                        "reason": f"Project {project_id} not found",
                        "data": row,
                    }
                )
                continue

            # ---- source_system default ----
            source_system = source_system_raw or "CALLING"

            # ---- external_id (ensure unique per (source_system, external_id)) ----
            safe_mobile = "".join(ch for ch in mobile if ch.isdigit()) or "NA"
            base_external_id = external_id_raw or f"CALL-{project_id}-{safe_mobile}-{row_no}"
            external_id = base_external_id
            suffix = 1

            # agar same source_system + external_id milta hai, to #1, #2 append kar do
            while LeadOpportunity.objects.filter(
                source_system=source_system,
                external_id=external_id,
            ).exists():
                external_id = f"{base_external_id}#{suffix}"
                suffix += 1

            # ---- owner resolve (username > email) ----
            owner = None
            owner_identifier = owner_username or owner_email

            if owner_identifier:
                try:
                    if owner_username:
                        owner = UserModel.objects.get(username__iexact=owner_username)
                    else:
                        owner = UserModel.objects.get(email__iexact=owner_email)
                except UserModel.DoesNotExist:
                    errors.append(
                        {
                            "row": row_no,
                            "reason": (
                                f'Owner user "{owner_identifier}" not found. '
                                "Owner left blank."
                            ),
                            "data": row,
                        }
                    )
                    owner = None
                except UserModel.MultipleObjectsReturned:
                    errors.append(
                        {
                            "row": row_no,
                            "reason": (
                                f'Owner user "{owner_identifier}" is not unique. '
                                "Owner left blank."
                            ),
                            "data": row,
                        }
                    )
                    owner = None

            # ---- to_lead flag ----
            to_lead = to_lead_raw in truthy_values

            # ==== DUPLICATE CHECK: same project + email + mobile ====
            dup_qs = LeadOpportunity.objects.filter(
                project_id=project.id,
                email__iexact=email,
                mobile_number=mobile,
            )
            existing_dup = dup_qs.order_by("-id").first()
            if existing_dup:
                skipped.append(
                    {
                        "row": row_no,
                        "opportunity_id": existing_dup.id,
                        "reason": "Duplicate (same project + email + mobile_number)",
                        "data": row,
                    }
                )
                continue

            # ==== CREATE NEW OPPORTUNITY + STATUS LOGIC ====
            try:
                with transaction.atomic():
                    create_kwargs = {
                        "source_system": source_system,
                        "external_id": external_id,
                        "project": project,
                        "source_name": source_name_raw,
                        "full_name": full_name,
                        "email": email,
                        "mobile_number": mobile,
                        "created_by": request.user
                        if request.user.is_authenticated
                        else None,
                    }
                    if owner is not None and hasattr(LeadOpportunity, "owner"):
                        create_kwargs["owner"] = owner

                    opp = LeadOpportunity.objects.create(**create_kwargs)

                    # ---- status_config / auto-convert ----
                    status_cfg = None
                    cfg_qs = LeadOpportunityStatusConfig.objects.for_project(
                        project.id
                    )

                    if status_code_raw:
                        status_cfg = cfg_qs.filter(
                            code=status_code_raw,
                            is_active=True,
                        ).first()

                    if not status_cfg and to_lead:
                        status_cfg = (
                            cfg_qs.filter(
                                code="CONVERTED",
                                can_convert=True,
                                is_active=True,
                            ).first()
                            or cfg_qs.filter(
                                can_convert=True,
                                is_active=True,
                            ).first()
                        )

                    auto_converted = False
                    sales_lead_id = None

                    if status_cfg:
                        payload = {
                            "status_config_id": status_cfg.id,
                            "comment": remark,
                        }
                        ser = LeadOpportunityStatusChangeSerializer(
                            data=payload,
                            context={"request": request, "opportunity": opp},
                        )
                        ser.is_valid(raise_exception=True)
                        result = ser.save()

                        auto_converted = bool(result.get("auto_converted"))
                        if result.get("sales_lead"):
                            sales_lead_id = result["sales_lead"].id

                    processed.append(
                        {
                            "row": row_no,
                            "opportunity_id": opp.id,
                            "created": True,
                            "status_config_id": status_cfg.id if status_cfg else None,
                            "status_config_code": status_cfg.code if status_cfg else None,
                            "status_config_label": status_cfg.label if status_cfg else None,
                            "auto_converted": auto_converted,
                            "sales_lead_id": sales_lead_id,
                        }
                    )

            except Exception as e:
                errors.append(
                    {
                        "row": row_no,
                        "reason": f"Exception: {e}",
                        "data": row,
                    }
                )
                continue

        total_rows = len(processed) + len(errors) + len(skipped)

        return Response(
            {
                "summary": {
                    "total_rows": total_rows,
                    "processed": len(processed),
                    "errors": len(errors),
                    "skipped": len(skipped),
                },
                "processed": processed,
                "errors": errors,
                "skipped": skipped,
            },
            status=status.HTTP_200_OK,
        )








class LeadOpportunityStatusConfigViewSet(viewsets.ReadOnlyModelViewSet):
    """
    List status configs for a project (project-specific + global).

    GET /api/sales/lead-opportunity-status-configs/?project_id=123
    """
    queryset = LeadOpportunityStatusConfig.objects.filter(is_active=True)
    serializer_class = LeadOpportunityStatusConfigSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        project_id = self.request.query_params.get("project_id")

        if project_id:
            qs = qs.filter(
                Q(project_id=project_id) | Q(project__isnull=True)
            ).order_by("-project_id", "code")
        else:
            # Agar project_id nahi diya to sirf GLOBAL dikhana
            qs = qs.filter(project__isnull=True).order_by("code")

        return qs






# salelead/views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied

from .models import PaymentLead, SalesLead
from .serializers import PaymentLeadSerializer,SiteVisitNORMALSerializer




class PaymentLeadViewSet(viewsets.ModelViewSet):
    """
    /api/sales/payment-leads/...

    - POST   /payment-leads/                      -> create payment for a lead
    - GET    /payment-leads/?lead=<id>            -> list payments filtered by lead
    - GET    /payment-leads/by-lead/<lead_id>/    -> list payments by lead (nice for FE)
    - GET    /payment-leads/<pk>/                 -> detail
    - PATCH  /payment-leads/<pk>/                 -> update
    - DELETE /payment-leads/<pk>/                 -> delete

    Extra:
    - GET    /payment-leads/pending/              -> all PENDING payments (accessible projects)
    - POST   /payment-leads/<id>/approve/         -> mark payment SUCCESS
    - POST   /payment-leads/<id>/reject/          -> mark payment FAILED / REFUNDED
    """

    serializer_class = PaymentLeadSerializer
    permission_classes = [IsAuthenticatedAndActive]

    # def get_queryset(self):
    #     user = self.request.user
    #     project_ids = _project_ids_for_user(user)

    #     qs = (
    #         PaymentLead.objects
    #         .select_related("lead", "project", "booking", "created_by")
    #         .filter(project_id__in=project_ids, for_kyc=False)
    #     )

    #     q = self.request.query_params

    #     # optional filter: ?lead=<lead_id>
    #     lead_id = q.get("lead")
    #     if lead_id:
    #         qs = qs.filter(lead_id=lead_id)

    #     # optional filter: ?payment_type=EOI / BOOKING
    #     ptype = q.get("payment_type")
    #     if ptype:
    #         qs = qs.filter(payment_type=ptype)

    #     # optional filter: ?status=PENDING / SUCCESS / ...
    #     status_val = q.get("status")
    #     if status_val:
    #         qs = qs.filter(status=status_val)

    #     # optional filter: ?project_id=<id>
    #     project_id = q.get("project_id")
    #     if project_id:
    #         qs = qs.filter(project_id=project_id)

    #     return qs.order_by("-payment_date", "-id")

    # def perform_create(self, serializer):
    #     """
    #     - project always lead.project se set hoga
    #     - created_by = request.user
    #     - security: user should have access to that project
    #     """
    #     lead = serializer.validated_data["lead"]

    #     user = self.request.user
    #     project_ids = _project_ids_for_user(user)
    #     if lead.project_id not in project_ids:
    #         raise PermissionDenied("You cannot add payment for this lead.")

    #     serializer.save(
    #         project=lead.project,
    #         created_by=user,
    # for_kyc=False,
    #     )

    def get_queryset(self):
        user = self.request.user
        project_ids = _project_ids_for_user(user)

        # ✅ IMPORTANT: prevent empty queryset for FULL_CONTROL
        if not project_ids:
            return PaymentLead.objects.none()

        qs = (
            PaymentLead.objects
            .select_related("lead", "project", "booking", "created_by")
            .filter(project_id__in=project_ids, for_kyc=False)
        )

        q = self.request.query_params

        if q.get("lead"):
            qs = qs.filter(lead_id=q["lead"])

        if q.get("payment_type"):
            qs = qs.filter(payment_type=q["payment_type"])

        if q.get("status"):
            qs = qs.filter(status=q["status"])

        if q.get("project_id"):
            qs = qs.filter(project_id=q["project_id"])

        return qs.order_by("-payment_date", "-id")
    
    def perform_create(self, serializer):
        lead = serializer.validated_data["lead"]
        user = self.request.user

        project_ids = _project_ids_for_user(user)
        if lead.project_id not in project_ids:
            raise PermissionDenied("You cannot add payment for this lead.")

        serializer.save(
            project=lead.project,
            created_by=user,
            for_kyc=False,
        )



    # ---------- BY LEAD (existing) ----------
        project_ids = _project_ids_for_user(user)
        if not project_ids:
         return Response([], status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r"by-lead/(?P<lead_id>[^/.]+)")
    def by_lead(self, request, lead_id=None):
        """
        GET /api/sales/payment-leads/by-lead/<lead_id>/

        Lead detail page ke liye:
        - saare payments (EOI + BOOKING)
        - sorted by latest first
        """
        user = request.user
        project_ids = _project_ids_for_user(user)

        try:
            lead = (
                SalesLead.objects
                .select_related("project")
                .get(pk=lead_id, project_id__in=project_ids)
            )
        except SalesLead.DoesNotExist:
            return Response(
                {"detail": "Lead not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        qs = (
            PaymentLead.objects
            .filter(lead=lead, for_kyc=False)
            .select_related("project", "booking", "created_by")
            .order_by("-payment_date", "-id")
        )

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ---------- NEW: PENDING LIST ----------
    @action(detail=False, methods=["get"], url_path="pending")
    def pending(self, request):
        """
        GET /api/sales/payment-leads/pending/?project_id=&lead=&payment_type=&payment_method=

        Returns all PENDING payments for projects jinke upar
        current user ka access hai.
        """
        user = request.user
        project_ids = _project_ids_for_user(user)

        q = request.query_params

        qs = (
            PaymentLead.objects
            .select_related("lead", "project", "booking", "created_by")
            .filter(project_id__in=project_ids, status=PaymentLead.PaymentStatus.PENDING, for_kyc=False)
        )

        if pid := q.get("project_id"):
            qs = qs.filter(project_id=pid)

        if lead_id := q.get("lead"):
            qs = qs.filter(lead_id=lead_id)

        if ptype := q.get("payment_type"):
            qs = qs.filter(payment_type=ptype)

        if pmethod := q.get("payment_method"):
            qs = qs.filter(payment_method=pmethod)

        qs = qs.order_by("-payment_date", "-id")

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ---------- helpers for approve / reject ----------
    def _check_payment_permission(self, payment, user):
        """
        Ensure user has project access. (same as list/create)
        """
        project_ids = _project_ids_for_user(user)
        if payment.project_id not in project_ids:
            raise PermissionDenied("You cannot modify payments for this project.")

    def _ensure_pending(self, payment):
        if payment.status != PaymentLead.PaymentStatus.PENDING:
            raise PermissionDenied("Only PENDING payments can be updated.")

    # ---------- APPROVE ----------
    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        """
        POST /api/sales/payment-leads/<id>/approve/

        Body (optional):
        {
          "reason": "Approved after checking bank statement"
        }

        Effect:
        - status: PENDING -> SUCCESS
        - notes: optional reason append
        """
        payment = self.get_object()
        user = request.user

        # Security
        self._check_payment_permission(payment, user)
        self._ensure_pending(payment)

        reason = (request.data.get("reason") or "").strip()

        # Optional: append reason into notes
        if reason:
            existing = (payment.notes or "").strip()
            if existing:
                payment.notes = f"{existing}\n[APPROVED] {reason}"
            else:
                payment.notes = f"[APPROVED] {reason}"

        payment.status = PaymentLead.PaymentStatus.SUCCESS
        payment.save(update_fields=["status", "notes"])

        serializer = self.get_serializer(payment)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ---------- REJECT ----------
    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        """
        POST /api/sales/payment-leads/<id>/reject/

        Body:
        {
          "reason": "Cheque bounced",
          "final_status": "FAILED" | "REFUNDED"   # optional, default = FAILED
        }

        Effect:
        - status: PENDING -> FAILED (default) or REFUNDED
        - notes: mark reason
        """
        payment = self.get_object()
        user = request.user

        # Security
        self._check_payment_permission(payment, user)
        self._ensure_pending(payment)

        reason = (request.data.get("reason") or "").strip()
        final_status = (request.data.get("final_status") or "").upper().strip()

        # Allow only FAILED / REFUNDED – anything else -> FAILED
        allowed = {
            PaymentLead.PaymentStatus.FAILED,
            PaymentLead.PaymentStatus.REFUNDED,
        }
        if final_status not in allowed:
            final_status = PaymentLead.PaymentStatus.FAILED

        # Append reason to notes
        if reason:
            existing = (payment.notes or "").strip()
            if existing:
                payment.notes = f"{existing}\n[REJECTED {final_status}] {reason}"
            else:
                payment.notes = f"[REJECTED {final_status}] {reason}"

        payment.status = final_status
        payment.save(update_fields=["status", "notes"])

        serializer = self.get_serializer(payment)
        return Response(serializer.data, status=status.HTTP_200_OK)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .serializers import (
    OnsiteRegistrationSerializer,
    SiteVisitNORMALSerializer,
)


class OnsiteRegistrationAPIView(APIView):
    """
    POST /api/sales/onsite-registration/
    """

    permission_classes = [IsAuthenticatedAndActive]

    def post(self, request, *args, **kwargs):
        serializer = OnsiteRegistrationSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        lead, stage_history, site_visit = serializer.save()

        lead_data = SalesLeadFullDetailSerializer(
            lead, context={"request": request}
        ).data

        visit_data = SiteVisitNORMALSerializer(
            site_visit, context={"request": request}
        ).data

        return Response(
            {
                "present": True,
                "lead": lead_data,
                "site_visit": visit_data,
                "stage_history_id": stage_history.id if stage_history else None,
            },
            status=status.HTTP_201_CREATED,
        )
